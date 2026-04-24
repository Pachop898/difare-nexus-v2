"""
agente/api_gerencial.py — Blueprint Flask con los endpoints del Dashboard Gerencial
====================================================================================
Endpoints expuestos (todos requieren JWT con rol 'admin' o 'gerencial'):

  GET  /api/kpis                → KPIs cabecera del dashboard
  GET  /api/tendencia-marca     → ?un=FARMACIAS|DIFARE&yoy=0|1
  GET  /api/venta-canal-mes     → Barras agrupadas por canal/mes
  GET  /api/ranking-pdv         → ?canal=FARMACIAS|DIFARE S.A.&top=50
  GET  /api/pareto-pdv          → Top farmacias que acumulan 80% de venta
  POST /api/chat-gerencial      → Chat con function calling (3 tools)

Diseñado para integrarse con app.py SIN modificarlo más allá del
register_blueprint en una línea. La verificación JWT reusa la función
`verificar_jwt` del módulo padre vía late-binding (set_jwt_verifier).
"""

from __future__ import annotations
import os, json, traceback
from flask import Blueprint, request, jsonify, send_file
from . import analitica

# ── Cliente Anthropic (singleton propio del blueprint) ──
_anthropic_client = None
_anthropic_client_fn = None  # fallback: función inyectada desde app.py

def set_anthropic_client(fn):
    """app.py pasa get_anthropic_client() como fallback."""
    global _anthropic_client_fn
    _anthropic_client_fn = fn

def _get_client():
    """Obtiene el cliente Anthropic, probando múltiples vías."""
    global _anthropic_client
    if _anthropic_client is not None:
        return _anthropic_client
    # 1) Intentar via función inyectada desde app.py
    if _anthropic_client_fn:
        try:
            c = _anthropic_client_fn()
            if c:
                _anthropic_client = c
                return c
        except Exception:
            pass
    # 2) Crear directamente (SDK auto-detecta ANTHROPIC_API_KEY del env)
    import anthropic
    key = os.getenv("ANTHROPIC_API_KEY", "")
    if key:
        _anthropic_client = anthropic.Anthropic(api_key=key)
    else:
        _anthropic_client = anthropic.Anthropic()  # intenta default
    return _anthropic_client

bp = Blueprint("gerencial", __name__, url_prefix="/api")

# Late-binding del verificador JWT y del mapa de roles ────────────────────
_jwt_verifier = None
_roles = {}  # { "francisco": "admin", "Gerente": "gerencial", "Campo": "campo" }

def set_jwt_verifier(fn, roles_map: dict):
    """app.py llama esto al registrar el blueprint."""
    global _jwt_verifier, _roles
    _jwt_verifier = fn
    _roles = roles_map


def _autorizar(roles_permitidos=("admin", "gerencial")):
    """Devuelve (usuario, rol) si OK; o (None, error_response) si no."""
    if _jwt_verifier is None:
        return None, (jsonify({"error": "JWT no configurado"}), 500)
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        token = request.args.get("token", "")
    usuario = _jwt_verifier(token)
    if not usuario:
        return None, (jsonify({"error": "No autorizado"}), 401)
    rol = _roles.get(usuario, "campo")
    if rol not in roles_permitidos:
        return None, (jsonify({"error": f"Rol '{rol}' no autorizado"}), 403)
    return (usuario, rol), None


# ══════════════════════════════════════════════════════════════
# 1) KPIs cabecera del dashboard
# ══════════════════════════════════════════════════════════════

def _parse_filtros():
    """Extrae filtros comunes de los query params."""
    marca = request.args.get("marca", "").strip() or None
    canal = request.args.get("canal", "").strip() or None
    grupos = request.args.getlist("grupo") or None
    productos = request.args.getlist("producto") or None
    if grupos and grupos == [""]:
        grupos = None
    if productos and productos == [""]:
        productos = None
    return marca, canal, grupos, productos


@bp.route("/kpis", methods=["GET"])
def kpis():
    auth, err = _autorizar()
    if err: return err
    try:
        marca, canal, grupos, productos = _parse_filtros()
        return jsonify(analitica.kpis_dashboard(
            marca=marca, canal=canal, grupos=grupos, productos=productos)), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/dois", methods=["GET"])
def dois():
    auth, err = _autorizar()
    if err: return err
    try:
        marca, canal, grupos, productos = _parse_filtros()
        producto = productos[0] if productos else None
        return jsonify(analitica.dias_inventario(
            producto=producto, marca=marca, canal=canal, grupos=grupos)), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/filtros", methods=["GET"])
def filtros():
    auth, err = _autorizar()
    if err: return err
    try:
        marca = request.args.get("marca", "").strip() or None
        return jsonify(analitica.filtros_disponibles(marca=marca)), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
# 2) Tendencia por marca (pregunta KAM #1)
# ══════════════════════════════════════════════════════════════

@bp.route("/tendencia-marca", methods=["GET"])
def tendencia_marca():
    auth, err = _autorizar()
    if err: return err
    un_param = request.args.get("un", "").strip().upper()
    un_map = {"FARMACIAS": "FARMACIAS",
              "DIFARE": "DISTRIBUCION DIFARE",
              "DISTRIBUCION": "DISTRIBUCION DIFARE",
              "": None, "TOTAL": None}
    unidad = un_map.get(un_param, None)
    yoy = request.args.get("yoy", "0") in ("1", "true", "True")
    try:
        data = analitica.tendencia_marca(unidad_negocio=unidad, comparar_yoy=yoy)
        return jsonify({
            "unidad_negocio": unidad or "TOTAL",
            "yoy": yoy,
            "filas": data,
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
# 2b) Venta por canal por mes (barras agrupadas dashboard)
# ══════════════════════════════════════════════════════════════

@bp.route("/venta-canal-mes", methods=["GET"])
def venta_canal_mes():
    auth, err = _autorizar()
    if err: return err
    try:
        marca, canal, grupos, productos = _parse_filtros()
        return jsonify({"filas": analitica.venta_por_canal_mes(
            marca=marca, canal=canal, grupos=grupos, productos=productos)}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
# 3) Ranking PDV (pregunta KAM #4)
# ══════════════════════════════════════════════════════════════

@bp.route("/ranking-pdv", methods=["GET"])
def ranking_pdv():
    auth, err = _autorizar()
    if err: return err
    canal = request.args.get("canal", "FARMACIAS").strip().upper()
    canal_map = {"FARMACIAS": "FARMACIAS",
                 "DIFARE": "DISTRIBUCION DIFARE",
                 "DISTRIBUCION": "DISTRIBUCION DIFARE"}
    canal_real = canal_map.get(canal, "FARMACIAS")
    try:
        top = int(request.args.get("top", "50"))
    except ValueError:
        top = 50
    try:
        data = analitica.ranking_pdv(canal=canal_real, top_n=top)
        return jsonify({
            "canal": canal_real,
            "top_n": top,
            "filas": data,
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
# 4) Pareto PDV (preview)
# ══════════════════════════════════════════════════════════════

@bp.route("/pareto-pdv", methods=["GET"])
def pareto_pdv():
    auth, err = _autorizar()
    if err: return err
    try:
        data = analitica.pareto_pdv()
        return jsonify({
            "total_pdv_pareto": len(data),
            "filas": data,
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
# 5) Tienda Perfecta — Pareto con buckets exclusivos de stock
# ══════════════════════════════════════════════════════════════

@bp.route("/tienda-perfecta", methods=["GET"])
def tienda_perfecta():
    auth, err = _autorizar()
    if err: return err
    try:
        marca, canal, grupos, productos = _parse_filtros()
        grupo = grupos[0] if grupos else None
        rows = analitica.oportunidad_vectorizacion(
            marca=marca, producto=productos if productos else None, grupo=grupo
        )
        # Calcular buckets EXCLUSIVOS de stock
        for r in rows:
            s0 = r.get("STOCK_0", 0) or 0
            s1 = r.get("STOCK_1", 0) or 0
            s2 = r.get("STOCK_2", 0) or 0
            s3 = r.get("STOCK_3", 0) or 0
            r["stock_solo_0"] = s0
            r["stock_solo_2"] = max(s2 - s1, 0)
            uni = r.get("UNIVERSO_PDV", 0) or 1
            pres = r.get("PDV_PRESENCIA", 0) or 0
            r["cobertura_pct"] = round(pres / uni * 100, 1)
            # %Pon (ponderada) = PDVs con venta en último mes completo / PDVs con presencia
            venta_ult = r.get("PDV_VENTA_ULT_MES", 0) or 0
            r["ponderada_pct"] = round(venta_ult / pres * 100, 1) if pres > 0 else 0.0
        # Obtener fecha del último día de stock
        from agente import generar_pdfs as gp
        try:
            _, ultimo_dia_stock, _, _ = gp.detectar_ultimo_dia_stock_y_venta()
        except Exception:
            ultimo_dia_stock = None
        return jsonify({"filas": rows, "ultimo_dia_stock": ultimo_dia_stock}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Tienda Perfecta — descarga directa Excel vectorización ──

@bp.route("/tienda-perfecta-excel", methods=["GET"])
def tienda_perfecta_excel():
    auth, err = _autorizar()
    if err: return err
    try:
        import tempfile
        from datetime import datetime
        marca, canal, grupos, productos = _parse_filtros()
        grupo = grupos[0] if grupos else None
        # productos: pasar lista completa al Excel (soporta multi-select)
        producto_filtro = productos if productos else ""
        # tipos_pdv: filtro de qué PDVs incluir en el Excel
        tipos_pdv = request.args.getlist("tipo_pdv") or None
        if tipos_pdv and tipos_pdv == [""]:
            tipos_pdv = None
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        fname = f"vectorizacion_completo_{ts}.xlsx"
        ruta = os.path.join(tempfile.gettempdir(), fname)
        analitica.exportar_vectorizacion_excel(
            producto=producto_filtro,
            ruta_salida=ruta,
            marca=marca or "",
            grupo=grupo or "",
            tipos_pdv=tipos_pdv
        )
        if not os.path.exists(ruta):
            return jsonify({"error": "No se pudo generar el archivo"}), 500
        return send_file(ruta, as_attachment=True, download_name=fname)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)[:300]}), 500


# ══════════════════════════════════════════════════════════════
# 6) Plan de Visibilidad — análisis InStore
# ══════════════════════════════════════════════════════════════

@bp.route("/visibilidad", methods=["GET"])
def plan_visibilidad():
    auth, err = _autorizar()
    if err: return err
    try:
        from agente.analitica_visibilidad import analisis_visibilidad
        result = analisis_visibilidad()
        return jsonify(result), 200
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
# 7) Distribución Numérica — clientes por mes con filtro marca
# ══════════════════════════════════════════════════════════════

@bp.route("/dist-numerica-chart", methods=["GET"])
def dist_numerica_chart():
    auth, err = _autorizar()
    if err: return err
    marca = request.args.get("marca", "").strip() or None
    try:
        data = analitica.distribucion_numerica(marca=marca, top_n=0)
        # También devolver lista de marcas disponibles para el filtro
        d = analitica.cargar_data()
        df = d["df_todos"]
        dist = df[df["UNIDAD"] == "DISTRIBUCION DIFARE"]
        marcas = sorted(dist["MARCA"].dropna().unique().tolist()) if not dist.empty else []
        return jsonify({
            "resumen_meses": data.get("resumen_meses", []),
            "total_clientes": data.get("total_clientes_historico", 0),
            "marcas_disponibles": marcas,
            "marca_filtro": marca,
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
# 7) Reporte PDF estructurado
# ══════════════════════════════════════════════════════════════

@bp.route("/reporte-pdf", methods=["GET"])
def reporte_pdf():
    auth, err = _autorizar()
    if err: return err
    try:
        import tempfile
        from datetime import datetime
        from openpyxl import Workbook  # solo para verificar dependencias

        kpis = analitica.kpis_dashboard()
        dois = analitica.dias_inventario()
        venta_mes = analitica.venta_por_canal_mes()
        tp = analitica.oportunidad_vectorizacion(top_n=30)

        # Generar PDF con fpdf2 (ligero, sin reportlab)
        try:
            from fpdf import FPDF
        except ImportError:
            # Fallback: generar HTML que se puede imprimir como PDF
            return _generar_reporte_html(kpis, dois, venta_mes, tp)

        # Helper: sanitizar unicode para Helvetica (no soporta em-dash, etc.)
        def _s(text):
            return str(text).replace("\u2014", "-").replace("\u2013", "-").replace("\u2019", "'").replace("\u201c", '"').replace("\u201d", '"').encode("latin-1", "replace").decode("latin-1")

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # Header
        pdf.set_fill_color(10, 22, 40)
        pdf.rect(0, 0, 210, 40, "F")
        pdf.set_text_color(201, 168, 76)
        pdf.set_font("Helvetica", "B", 22)
        pdf.cell(0, 20, "ORION", ln=True, align="C")
        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(122, 143, 187)
        pdf.cell(0, 8, f"Reporte Gerencial - {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align="C")
        pdf.ln(10)

        # KPIs
        pdf.set_text_color(201, 168, 76)
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "KPIs Principales", ln=True)
        pdf.set_text_color(40, 40, 40)
        pdf.set_font("Helvetica", "", 11)
        vt = kpis.get("venta_total", 0)
        vf = kpis.get("venta_farmacias", 0)
        vd = kpis.get("venta_distribucion", 0)
        pdf.cell(0, 7, f"Venta Total: ${vt:,.2f}", ln=True)
        pdf.cell(0, 7, f"Farmacias: ${vf:,.2f} ({round(vf/max(vt,1)*100)}%)", ln=True)
        pdf.cell(0, 7, f"Distribucion: ${vd:,.2f} ({round(vd/max(vt,1)*100)}%)", ln=True)
        pdf.cell(0, 7, f"Universo PDV: {kpis.get('universo_pdv', 0)}", ln=True)
        dia = kpis.get("ultimo_dia_venta", "?")
        dias_m = kpis.get("dias_mes", "?")
        pdf.cell(0, 7, f"Data hasta dia {dia}/{dias_m} de abril 2026", ln=True)
        pdf.ln(6)

        # DOIS
        pdf.set_text_color(201, 168, 76)
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Dias de Inventario (DOIS)", ln=True)
        pdf.set_text_color(40, 40, 40)
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 7, f"DOIS Bodega: {dois.get('dois_bodega', '?')} dias", ln=True)
        pdf.cell(0, 7, f"DOIS PDV: {dois.get('dois_pdv', '?')} dias", ln=True)
        pdf.cell(0, 7, f"DOIS Total: {dois.get('dois_total', '?')} dias", ln=True)
        pdf.cell(0, 7, _s(f"Estado: {dois.get('estado_inventario', '?')}"), ln=True)
        pdf.ln(6)

        # Venta mensual
        pdf.set_text_color(201, 168, 76)
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Venta Mensual por Canal", ln=True)
        meses_nombre = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
        # Table header
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(27, 58, 107)
        pdf.set_text_color(201, 168, 76)
        pdf.cell(30, 8, "Mes", 1, 0, "C", True)
        pdf.cell(45, 8, "Farmacias", 1, 0, "C", True)
        pdf.cell(45, 8, "Distribucion", 1, 0, "C", True)
        pdf.cell(45, 8, "Total", 1, 1, "C", True)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(40, 40, 40)
        for fila in venta_mes:
            m = fila.get("mes", 0)
            nombre = meses_nombre[m] if 0 < m < 13 else str(m)
            if fila.get("proyectado"):
                nombre += " *"
            pdf.cell(30, 7, nombre, 1, 0, "C")
            pdf.cell(45, 7, f"${fila.get('farmacias', 0):,.0f}", 1, 0, "R")
            pdf.cell(45, 7, f"${fila.get('distribucion', 0):,.0f}", 1, 0, "R")
            pdf.cell(45, 7, f"${fila.get('total', 0):,.0f}", 1, 1, "R")
        pdf.ln(6)

        # Tienda Perfecta top 15
        pdf.set_text_color(201, 168, 76)
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Top 15 Productos Pareto - Tienda Perfecta", ln=True)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(27, 58, 107)
        pdf.set_text_color(201, 168, 76)
        pdf.cell(30, 7, "Marca", 1, 0, "C", True)
        pdf.cell(55, 7, "Producto", 1, 0, "C", True)
        pdf.cell(25, 7, "Venta", 1, 0, "C", True)
        pdf.cell(15, 7, "%Cob", 1, 0, "C", True)
        pdf.cell(20, 7, "Stock=0", 1, 0, "C", True)
        pdf.cell(20, 7, "Stock=1", 1, 0, "C", True)
        pdf.cell(20, 7, "Stock=2", 1, 1, "C", True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(40, 40, 40)
        for r in tp[:15]:
            uni = r.get("UNIVERSO_PDV", 0) or 1
            pres = r.get("PDV_PRESENCIA", 0)
            cob = round(pres / uni * 100, 1)
            s0 = r.get("STOCK_0", 0)
            s1 = max((r.get("STOCK_1", 0) or 0) - (s0 or 0), 0)
            s2 = max((r.get("STOCK_2", 0) or 0) - (r.get("STOCK_1", 0) or 0), 0)
            pdf.cell(30, 6, _s(str(r.get("MARCA", ""))[:15]), 1, 0, "L")
            pdf.cell(55, 6, _s(str(r.get("PRODUCTO", ""))[:30]), 1, 0, "L")
            pdf.cell(25, 6, f"${r.get('VENTA', r.get('venta_total', 0)):,.0f}", 1, 0, "R")
            pdf.cell(15, 6, f"{cob}%", 1, 0, "C")
            pdf.cell(20, 6, str(s0), 1, 0, "C")
            pdf.cell(20, 6, str(s1), 1, 0, "C")
            pdf.cell(20, 6, str(s2), 1, 1, "C")

        # Footer
        pdf.ln(10)
        pdf.set_text_color(122, 143, 187)
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 7, "Generado automaticamente por ORION - Genommalab Ecuador", ln=True, align="C")

        ts = datetime.now().strftime("%Y%m%d_%H%M")
        fname = f"reporte_nexus_{ts}.pdf"
        ruta = os.path.join(tempfile.gettempdir(), fname)
        pdf.output(ruta)
        return send_file(ruta, as_attachment=True, download_name=fname)

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)[:300]}), 500


def _generar_reporte_html(kpis, dois, venta_mes, tp):
    """Fallback: genera HTML imprimible si fpdf2 no está disponible."""
    from flask import Response
    vt = kpis.get("venta_total", 0)
    vf = kpis.get("venta_farmacias", 0)
    vd = kpis.get("venta_distribucion", 0)
    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
    <title>Reporte Nexus</title>
    <style>body{{font-family:Arial;max-width:800px;margin:auto;padding:20px}}
    h1{{color:#C9A84C}}h2{{color:#1B3A6B;border-bottom:2px solid #C9A84C;padding-bottom:4px}}
    table{{border-collapse:collapse;width:100%;margin:12px 0}}th,td{{border:1px solid #ccc;padding:6px 10px;font-size:12px}}
    th{{background:#1B3A6B;color:#C9A84C}}
    @media print{{body{{margin:0}}}}
    </style></head><body>
    <h1>ORION - Reporte Gerencial</h1>
    <h2>KPIs</h2>
    <p>Venta Total: <b>${vt:,.2f}</b> | Farmacias: <b>${vf:,.2f}</b> | Distribución: <b>${vd:,.2f}</b></p>
    <h2>Días de Inventario</h2>
    <p>Bodega: <b>{dois.get('dois_bodega','?')}</b> | PDV: <b>{dois.get('dois_pdv','?')}</b> | Total: <b>{dois.get('dois_total','?')}</b> días</p>
    <p><i>{dois.get('estado_inventario','')}</i></p>
    <h2>Venta Mensual</h2>
    <table><tr><th>Mes</th><th>Farmacias</th><th>Distribución</th><th>Total</th></tr>"""
    meses_n = ["","Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    for f in venta_mes:
        m = f.get("mes", 0)
        html += f"<tr><td>{meses_n[m] if 0<m<13 else m}</td><td>${f.get('farmacias',0):,.0f}</td><td>${f.get('distribucion',0):,.0f}</td><td>${f.get('total',0):,.0f}</td></tr>"
    html += "</table><p style='color:#999;font-size:10px;text-align:center'>Usa Ctrl+P para imprimir como PDF</p></body></html>"
    return Response(html, mimetype="text/html")


# ══════════════════════════════════════════════════════════════
# Util: invalidar cache (admin only)
# ══════════════════════════════════════════════════════════════

@bp.route("/recargar-data", methods=["POST"])
def recargar_data():
    auth, err = _autorizar(roles_permitidos=("admin",))
    if err: return err
    analitica.invalidar_cache()
    return jsonify({"ok": True, "mensaje": "Cache invalidado, próxima consulta recargará desde excels/"}), 200


# ══════════════════════════════════════════════════════════════
# CHAT GERENCIAL — Claude con function calling
# ══════════════════════════════════════════════════════════════

_SYSTEM_GERENCIAL = """Eres el asistente analítico ORION para el equipo gerencial de Genommalab Ecuador.
Tu rol es responder preguntas estratégicas sobre ventas, inventario, tendencias de marca y
rendimiento de puntos de venta (PDV) usando las herramientas disponibles.

REGLAS:
- Siempre usa las herramientas (tools) para obtener datos REALES antes de responder.
- Responde en español ecuatoriano, tono profesional y conciso.
- Formatea montos en USD con separador de miles y 2 decimales: $1.234,56
- Cuando muestres tablas, usa formato markdown con | columna | columna |.
- Si el usuario pide exportar a Excel, usa la herramienta exportar_excel.
- Incluye insights accionables: no solo datos, también recomendaciones.
- Máximo 400 palabras por respuesta.

CONTEXTO DE NEGOCIO:
- Canales: FARMACIAS (sell-out en PDV propios de DIFARE), DISTRIBUCION DIFARE (sell-in a clientes externos)
- DIFARE S.A. = bodega central (solo stock, no venta directa)
- DOIS (Días de Inventario) se calcula VALORIZADO (USD): Stock_Valorizado / Venta_Diaria_Promedio_USD
- Umbrales DOIS: >30 días = OK, 15-30 = Stock Bajo, <15 = Crítico
- NO confundir unidades con valores. El inventario se mide en USD, no en unidades físicas.
- Meses disponibles: Ene-Mar 2026 (cerrados) + Abr 2026 (parcial, data semanal SAP)
- Canal de venta en FARMACIAS: la columna CANAL segrega la venta por: APP, Call Center, Mostrador, Tienda Virtual, WhatsApp, etc.
  * Mostrador = canal PRESENCIAL = canal FÍSICO
  * Todo lo demás = canal NO PRESENCIAL = canal DIGITAL = e-commerce
  * SÍ tenemos esta data. SIEMPRE usa venta_por_canal_farmacia para estas preguntas.
- Sinónimos del usuario: "canal digital" = "canal no presencial" = "e-commerce" = "venta online" = "venta en línea"
- Sinónimos del usuario: "canal físico" = "canal presencial" = "mostrador"

VOCABULARIO CLAVE DEL USUARIO:
- "Vectorización" u "Oportunidades de vectorización" = se refiere SIEMPRE a FARMACIAS PROPIAS.
  Significa identificar qué productos Pareto (80% de la venta) faltan en qué PDV (farmacias) y
  deberían estar presentes. Usa la herramienta oportunidad_vectorizacion para obtener el análisis
  Pareto con cobertura, presencia y stock por PDV. Muestra: producto, venta, cobertura actual,
  PDVs sin stock, y recomienda cuáles priorizar.

- "Distribución numérica" = se refiere al CANAL DISTRIBUTIVO (DISTRIBUCION DIFARE).
  Significa analizar cuántos clientes (por RUC) se han atendido, cuántos son nuevos vs perdidos
  mes a mes, y si cada cliente está comprando el portafolio TOP completo.
  Si el usuario pregunta por distribución numérica, primero PREGUNTA si se refiere al canal
  distributivo o a farmacias. Si confirma distributivo, usa la herramienta distribucion_numerica.
  Lo mínimo esperado es que cada cliente con RUC compre el portafolio TOP cada mes.
"""

_TOOLS_GERENCIAL = [
    {
        "name": "tendencia_marca",
        "description": "Obtiene la venta mensual desglosada por MARCA. Útil para ver evolución de cada marca mes a mes, detectar caídas o crecimientos. Puedes filtrar por canal (farmacias o distribución).",
        "input_schema": {
            "type": "object",
            "properties": {
                "canal": {
                    "type": "string",
                    "description": "Canal de venta: 'FARMACIAS', 'DISTRIBUCION DIFARE', o null para total.",
                    "enum": ["FARMACIAS", "DISTRIBUCION DIFARE", None]
                }
            },
            "required": []
        }
    },
    {
        "name": "dias_inventario",
        "description": "Calcula DOIS (Días de Inventario) VALORIZADO en USD: Stock_Valorizado / Venta_Diaria_Promedio_USD. Devuelve DOIS bodega, PDV y total. Umbrales: >30 días = OK, 15-30 = Stock Bajo, <15 = Crítico. NO usa unidades, usa valores monetarios.",
        "input_schema": {
            "type": "object",
            "properties": {
                "producto": {
                    "type": "string",
                    "description": "Nombre o código del producto a consultar. Dejar vacío para el total de todos los productos."
                }
            },
            "required": []
        }
    },
    {
        "name": "pareto_pdv",
        "description": "Lista las farmacias que concentran el 80% de la venta total (análisis Pareto/80-20). Muestra cada farmacia con su venta y porcentaje acumulado. Útil para priorizar visitas y negociaciones.",
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": []
        }
    },
    {
        "name": "ranking_pdv",
        "description": "Ranking de los mejores PDV o clientes por volumen de venta. Para farmacias muestra el Top N de PDV individuales. Para distribución muestra los clientes más grandes.",
        "input_schema": {
            "type": "object",
            "properties": {
                "canal": {
                    "type": "string",
                    "description": "Canal: 'FARMACIAS' o 'DISTRIBUCION DIFARE'.",
                    "enum": ["FARMACIAS", "DISTRIBUCION DIFARE"]
                },
                "top_n": {
                    "type": "integer",
                    "description": "Cantidad de resultados. Default: 20.",
                    "default": 20
                }
            },
            "required": []
        }
    },
    {
        "name": "kpis_resumen",
        "description": "Obtiene los KPIs principales: venta total, venta farmacias, venta distribución, universo de PDV, día de data, y stock valorizado por mes. Útil para contexto general antes de profundizar.",
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": []
        }
    },
    {
        "name": "venta_canal_mes",
        "description": "Venta mensual desglosada por canal (farmacias vs distribución) con proyección del mes en curso. Muestra barras por mes con el total de cada canal.",
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": []
        }
    },
    {
        "name": "exportar_excel",
        "description": "Genera el Informe de Vectorización semanal en Excel (.xlsx) para FARMACIAS. Contiene una pestaña por marca con los productos Pareto (80% de la venta). Cada fila es un PDV que necesita stock: sin vectorizar, stock=0, stock bajo. Incluye columna SUGERIDO con unidades mínimas a enviar. Usa esta herramienta cuando el usuario pide 'reporte de vectorización', 'informe semanal', 'exportar vectorización', o 'generar Excel para el cliente'. Si se pasa un producto específico, solo incluye ese producto.",
        "input_schema": {
            "type": "object",
            "properties": {
                "producto": {
                    "type": "string",
                    "description": "Producto o marca específica para filtrar (ej: 'NIKZON', 'SUEROX'). Dejar vacío para generar el informe completo con TODAS las marcas Pareto."
                }
            },
            "required": []
        }
    },
    {
        "name": "oportunidad_vectorizacion",
        "description": "Análisis de oportunidades de VECTORIZACIÓN en FARMACIAS propias. Devuelve los productos Pareto (80% de la venta) con: cobertura actual (% de PDVs que lo tienen), PDVs con stock=0, y presencia. Útil para identificar dónde faltan productos estrella y priorizar envíos. Usa esta herramienta cuando el usuario pregunta por 'vectorización', 'oportunidades', 'cobertura de productos', o 'qué falta en las farmacias'.",
        "input_schema": {
            "type": "object",
            "properties": {
                "producto": {
                    "type": "string",
                    "description": "Nombre del producto para filtrar. Dejar vacío para ver todos los productos Pareto."
                },
                "top_n": {
                    "type": "integer",
                    "description": "Cantidad de productos a mostrar. Default: 20.",
                    "default": 20
                }
            },
            "required": []
        }
    },
    {
        "name": "dois_por_producto",
        "description": "Lista el DOIS (Días de Inventario Valorizado) de CADA producto Pareto individualmente. Permite filtrar por rango: solo productos con DOIS mayor a X días, o menor a Y días. Útil para identificar productos con exceso de inventario (DOIS alto → necesitan promoción) o riesgo de quiebre (DOIS bajo → reabastecer). Devuelve: producto, marca, stock valorizado (bodega+PDV), venta diaria y DOIS.",
        "input_schema": {
            "type": "object",
            "properties": {
                "umbral_min": {
                    "type": "number",
                    "description": "DOIS mínimo para filtrar (ej: 90 para ver solo productos con >90 días de inventario). Default: 0."
                },
                "umbral_max": {
                    "type": "number",
                    "description": "DOIS máximo para filtrar (ej: 15 para ver solo productos en riesgo). Default: 9999."
                },
                "top_n": {
                    "type": "integer",
                    "description": "Cantidad máxima de resultados. Default: 30."
                }
            },
            "required": []
        }
    },
    {
        "name": "venta_por_canal_farmacia",
        "description": "Analiza la venta en FARMACIAS desglosada por CANAL de venta (columna CANAL en la data). Canales: Mostrador, APP, Call Center, Tienda Virtual, WhatsApp, etc. Clasificación: Mostrador = canal PRESENCIAL / FÍSICO. Todo lo demás = canal NO PRESENCIAL / DIGITAL / e-commerce. Muestra cuánto se vende por cada canal, el split presencial vs no presencial/digital, y la evolución mensual. IMPORTANTE: SIEMPRE usa esta herramienta cuando el usuario mencione CUALQUIERA de estos términos: 'canal digital', 'canal físico', 'canal presencial', 'canal no presencial', 'e-commerce', 'venta online', 'venta digital', 'venta presencial', 'venta por APP', 'mostrador', 'cómo se vende', 'por qué medio se vende', 'canales de venta', 'tipo de venta', 'medio de venta', 'venta en línea'. La data SÍ tiene esta información en la columna CANAL.",
        "input_schema": {
            "type": "object",
            "properties": {
                "marca": {
                    "type": "string",
                    "description": "Marca para filtrar (ej: 'CICATRICURE'). Dejar vacío para analizar todas."
                },
                "producto": {
                    "type": "string",
                    "description": "Producto específico para filtrar. Dejar vacío para todos."
                },
                "mes": {
                    "type": "integer",
                    "description": "Mes específico (1-12) para filtrar. Dejar vacío para ver todos los meses."
                }
            },
            "required": []
        }
    },
    {
        "name": "distribucion_numerica",
        "description": "Análisis de distribución numérica del canal DISTRIBUTIVO (DISTRIBUCION DIFARE): clientes atendidos por RUC mes a mes, clientes nuevos vs perdidos, y penetración del portafolio TOP. Usa esta herramienta cuando el usuario pregunta por 'distribución numérica', 'clientes atendidos', 'cuántos RUCs compraron', o 'penetración del portafolio'.",
        "input_schema": {
            "type": "object",
            "properties": {
                "marca": {
                    "type": "string",
                    "description": "Marca para filtrar (ej: 'CICATRICURE'). Dejar vacío para analizar todas las marcas."
                },
                "top_n": {
                    "type": "integer",
                    "description": "Cantidad de clientes a mostrar en penetración. Default: 20.",
                    "default": 20
                }
            },
            "required": []
        }
    },
]


def _ejecutar_tool(name: str, inp: dict) -> str:
    """Ejecuta una tool y devuelve resultado como string JSON."""
    try:
        if name == "tendencia_marca":
            data = analitica.tendencia_marca(
                unidad_negocio=inp.get("canal"),
                comparar_yoy=False
            )
            # Agrupar por marca para que sea más legible
            marcas = {}
            for r in data:
                m = r.get("marca", "?")
                marcas.setdefault(m, []).append(r)
            return json.dumps({"total_registros": len(data), "por_marca": marcas}, default=str, ensure_ascii=False)

        elif name == "dias_inventario":
            data = analitica.dias_inventario(producto=inp.get("producto") or None)
            return json.dumps(data, default=str, ensure_ascii=False)

        elif name == "pareto_pdv":
            data = analitica.pareto_pdv()
            return json.dumps({"total_pdv_80pct": len(data), "top_20": data[:20]}, default=str, ensure_ascii=False)

        elif name == "ranking_pdv":
            canal = inp.get("canal", "FARMACIAS")
            top_n = inp.get("top_n", 20)
            data = analitica.ranking_pdv(canal=canal, top_n=top_n)
            return json.dumps({"canal": canal, "filas": data}, default=str, ensure_ascii=False)

        elif name == "kpis_resumen":
            data = analitica.kpis_dashboard()
            return json.dumps(data, default=str, ensure_ascii=False)

        elif name == "venta_canal_mes":
            data = analitica.venta_por_canal_mes()
            return json.dumps({"filas": data}, default=str, ensure_ascii=False)

        elif name == "exportar_excel":
            producto = inp.get("producto", "")
            import tempfile
            from datetime import datetime
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            fname = f"vectorizacion_{producto.replace(' ','_') if producto else 'completo'}_{ts}.xlsx"
            ruta = os.path.join(tempfile.gettempdir(), fname)
            analitica.exportar_vectorizacion_excel(producto=producto, ruta_salida=ruta)
            return json.dumps({"ok": True, "ruta": ruta, "producto": producto or "TODAS LAS MARCAS",
                               "archivo": fname})

        elif name == "oportunidad_vectorizacion":
            data = analitica.oportunidad_vectorizacion(
                producto=inp.get("producto") or None,
                top_n=inp.get("top_n", 20)
            )
            # Enriquecer con % cobertura para que Claude lo presente mejor
            for r in data:
                uni = r.get("UNIVERSO_PDV", 0) or 1
                pres = r.get("PDV_PRESENCIA", 0)
                r["cobertura_pct"] = round(pres / uni * 100, 1)
                r["pdv_sin_stock"] = r.get("STOCK_0", 0)
                r["pdv_stock_critico"] = r.get("STOCK_1", 0)
            return json.dumps({
                "total_productos_pareto": len(data),
                "productos": data
            }, default=str, ensure_ascii=False)

        elif name == "dois_por_producto":
            data = analitica.dois_por_producto(
                umbral_min=inp.get("umbral_min", 0),
                umbral_max=inp.get("umbral_max", 9999),
                top_n=inp.get("top_n", 30),
            )
            return json.dumps({
                "total_productos": len(data),
                "productos": data,
            }, default=str, ensure_ascii=False)

        elif name == "venta_por_canal_farmacia":
            data = analitica.venta_por_canal_farmacia(
                marca=inp.get("marca") or None,
                producto=inp.get("producto") or None,
                mes=inp.get("mes") or None,
            )
            return json.dumps(data, default=str, ensure_ascii=False)

        elif name == "distribucion_numerica":
            data = analitica.distribucion_numerica(
                marca=inp.get("marca") or None,
                top_n=inp.get("top_n", 20)
            )
            return json.dumps(data, default=str, ensure_ascii=False)

        else:
            return json.dumps({"error": f"Tool desconocida: {name}"})
    except Exception as e:
        return json.dumps({"error": str(e)[:200]})


@bp.route("/chat-gerencial", methods=["POST", "OPTIONS"])
def chat_gerencial():
    if request.method == "OPTIONS":
        return "", 204
    auth, err = _autorizar()
    if err:
        return err

    body = request.get_json(silent=True) or {}
    pregunta = body.get("pregunta", "").strip()
    historial = body.get("historial", [])  # [{role, content}, ...]
    if not pregunta:
        return jsonify({"error": "Falta 'pregunta'"}), 400

    try:
        client = _get_client()
        # Verificar que la key tenga formato correcto
        key_val = os.getenv("ANTHROPIC_API_KEY", "")
        if not key_val:
            return jsonify({"error": "ANTHROPIC_API_KEY no está configurada en las variables de entorno del servidor."}), 500
        if not key_val.startswith("sk-ant-"):
            return jsonify({"error": f"ANTHROPIC_API_KEY tiene formato incorrecto (empieza con '{key_val[:8]}...'). Debe empezar con 'sk-ant-'."}), 500
    except Exception as e:
        return jsonify({"error": f"API key no configurada: {e}"}), 500

    # Construir mensajes
    messages = []
    for h in historial[-10:]:  # últimos 10 turnos de contexto
        messages.append({"role": h.get("role", "user"), "content": h.get("content", "")})
    messages.append({"role": "user", "content": pregunta})

    # Agentic loop: Claude puede llamar tools múltiples veces
    file_downloads = []  # archivos generados para descarga
    max_iterations = 5
    try:
        for _ in range(max_iterations):
            resp = client.messages.create(
                model="claude-sonnet-4-5",
                max_tokens=1024,
                system=_SYSTEM_GERENCIAL,
                tools=_TOOLS_GERENCIAL,
                messages=messages,
            )
            # Procesar respuesta
            if resp.stop_reason == "tool_use":
                # Claude quiere llamar una o más tools
                tool_results = []
                for block in resp.content:
                    if block.type == "tool_use":
                        result_str = _ejecutar_tool(block.name, block.input)
                        # Si fue exportar_excel, guardar ruta del archivo
                        if block.name == "exportar_excel":
                            try:
                                r = json.loads(result_str)
                                if r.get("ok") and r.get("ruta"):
                                    file_downloads.append(r["ruta"])
                            except Exception:
                                pass
                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block.id,
                            "content": result_str,
                        })
                # Agregar turno del asistente + resultados
                messages.append({"role": "assistant", "content": resp.content})
                messages.append({"role": "user", "content": tool_results})
            else:
                # Respuesta final
                text_parts = [b.text for b in resp.content if hasattr(b, "text")]
                respuesta = "\n".join(text_parts)
                result = {
                    "respuesta": respuesta,
                    "archivos": [os.path.basename(f) for f in file_downloads],
                }
                return jsonify(result), 200

        # Si agotamos iteraciones, devolver lo que haya
        text_parts = [b.text for b in resp.content if hasattr(b, "text")]
        fallback = "\n".join(text_parts) or "No pude completar el análisis con las herramientas disponibles. Intenta reformular la pregunta o ser más específico (ej: menciona una marca o producto concreto)."
        return jsonify({"respuesta": fallback, "archivos": []}), 200

    except Exception as e:
        traceback.print_exc()
        err_msg = str(e)
        if "authentication_error" in err_msg or "invalid x-api-key" in err_msg:
            return jsonify({"error": "La API key de Anthropic es inválida. Ve a console.anthropic.com > API Keys, genera una nueva key que empiece con 'sk-ant-api03-...' y actualízala en Railway."}), 500
        return jsonify({"error": f"Error en análisis: {err_msg[:200]}"}), 500


# Endpoint para descargar archivos generados por el chat
@bp.route("/descargar/<filename>", methods=["GET"])
def descargar_archivo(filename):
    # Aceptar token en header O en query param (para links <a href>)
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        token = request.args.get("token", "")
    if _jwt_verifier and token:
        usuario = _jwt_verifier(token)
        if not usuario:
            return jsonify({"error": "No autorizado"}), 401
    elif _jwt_verifier:
        return jsonify({"error": "No autorizado"}), 401
    import tempfile
    ruta = os.path.join(tempfile.gettempdir(), filename)
    if not os.path.exists(ruta):
        return jsonify({"error": "Archivo no encontrado"}), 404
    return send_file(ruta, as_attachment=True, download_name=filename)

"""
agente/analitica.py — Capa de cálculo pura para DIFARE NEXUS v2

Reutiliza las funciones ya validadas en generar_pdfs.py (proyecto agente-excel),
pero las expone como una API limpia que el backend Flask puede consumir desde
los endpoints /api/* sin tocar matplotlib ni reportlab.

Toda la lógica numérica vive en generar_pdfs.py (fuente de verdad). Aquí solo
re-exportamos y añadimos los wrappers nuevos que pide el dashboard gerencial:

  - tendencia_marca(unidad_negocio=None, comparar_yoy=False)
  - ranking_pdv(canal="FARMACIAS", top_n=50)
  - pareto_pdv()
  - dias_inventario(producto=None)
  - proyeccion_venta(horizonte_dias=30)
  - oportunidad_vectorizacion(producto)
  - sugerido_stock(grupo_farmacia=None, productos_top=20,
                   lead_time=2, buffer=8)
  - exportar_vectorizacion_excel(producto, ruta_salida)

Parámetros de negocio por defecto (Fase 1, confirmados con Pacho 2026-04-06):
  - LEAD_TIME_DIAS = 2
  - BUFFER_DIAS    = 8
  - DIAS_INV_SEGURIDAD = LEAD_TIME_DIAS + BUFFER_DIAS = 10
"""

from __future__ import annotations

import os
import sys
import pandas as pd

# generar_pdfs.py vive en la misma carpeta agente/. Lo importamos como
# módulo de cálculo. La carga de matplotlib en él es perezosa al renderizar,
# así que importarlo NO renderiza nada.
_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

# Stubs perezosos: generar_pdfs.py importa matplotlib y reportlab al cargar,
# pero analitica.py NUNCA renderiza gráficos ni PDFs. En entornos donde
# matplotlib/reportlab no están instalados (p. ej. tests rápidos del backend
# o un Railway sin matplotlib) inyectamos stubs vacíos para que el import
# no falle. La generación real de PDFs se hace en otra ruta y allí sí
# tendrá las dependencias instaladas.
import importlib, types
def _stub(name, attrs=None):
    if name in sys.modules:
        return
    m = types.ModuleType(name)
    for k, v in (attrs or {}).items():
        setattr(m, k, v)
    sys.modules[name] = m

try:
    importlib.import_module("matplotlib")
except ImportError:
    _stub("matplotlib")
    _stub("matplotlib.pyplot")
try:
    importlib.import_module("reportlab")
except ImportError:
    for n in ("reportlab", "reportlab.lib", "reportlab.lib.pagesizes",
              "reportlab.lib.colors", "reportlab.lib.units",
              "reportlab.platypus", "reportlab.lib.styles",
              "reportlab.lib.enums"):
        _stub(n)
    # atributos mínimos usados a nivel de módulo en generar_pdfs
    sys.modules["reportlab.lib.pagesizes"].A4 = None
    sys.modules["reportlab.lib.pagesizes"].landscape = lambda x: x
    class _ColorsStub:
        @staticmethod
        def HexColor(s): return s
        white = "#FFFFFF"
    sys.modules["reportlab.lib"].colors = _ColorsStub
    sys.modules["reportlab.lib.colors"] = _ColorsStub  # type: ignore
    sys.modules["reportlab.lib.units"].cm = 1
    for cls in ("SimpleDocTemplate", "Paragraph", "Spacer", "Table", "TableStyle", "Image"):
        setattr(sys.modules["reportlab.platypus"], cls, type(cls, (), {}))
    sys.modules["reportlab.lib.styles"].getSampleStyleSheet = lambda: {}
    sys.modules["reportlab.lib.styles"].ParagraphStyle = type("ParagraphStyle", (), {})
    sys.modules["reportlab.lib.enums"].TA_CENTER = 0
    sys.modules["reportlab.lib.enums"].TA_LEFT = 1

import generar_pdfs as gp  # noqa: E402

# ── Parámetros de negocio (configurables por panel admin en Fase 1) ──
LEAD_TIME_DIAS = 2
BUFFER_DIAS = 8
DIAS_INV_SEGURIDAD = LEAD_TIME_DIAS + BUFFER_DIAS  # 10

# Buscar carpeta excels/ en varias ubicaciones candidatas
_EXCELS_CANDIDATES = [
    os.path.join(os.path.dirname(_HERE), "excels"),              # /app/excels (Railway)
    os.path.join(os.path.dirname(os.path.dirname(_HERE)), "excels"),  # legacy
    os.path.join(_HERE, "excels"),                               # /app/agente/excels
    "excels",                                                     # relativo al CWD
]
EXCELS_DIR = next((p for p in _EXCELS_CANDIDATES if os.path.isdir(p)), _EXCELS_CANDIDATES[0])
print(f"[analitica] EXCELS_DIR = {EXCELS_DIR} (existe={os.path.isdir(EXCELS_DIR)})")


# ══════════════════════════════════════════════════════════════
# Carga base (cacheada en memoria por proceso)
# ══════════════════════════════════════════════════════════════

_cache = {}
_cache_ts = 0  # timestamp de última carga
_CACHE_TTL = 43200  # 12 horas — los Excel no cambian durante el día
import time as _time
import threading as _threading
# Lock global para que solo UN hilo haga la carga pesada a la vez.
# Evita que el pre-warm y un request concurrente ambos lean los Excel
# en paralelo → doble memoria → riesgo OOM en Railway.
_cargar_lock = _threading.Lock()

def _carpeta():
    return EXCELS_DIR if os.path.isdir(EXCELS_DIR) else "excels"

def _excels_mtime() -> float:
    """Retorna el mtime más reciente de cualquier .xlsx en la carpeta."""
    import glob
    carpeta = _carpeta()
    archivos = glob.glob(os.path.join(carpeta, "*.xlsx"))
    if not archivos:
        return 0
    return max(os.path.getmtime(f) for f in archivos)

def cargar_data(force: bool = False) -> dict:
    """
    Devuelve un dict con todos los DataFrames base ya calculados.
    Cachea en memoria con TTL de 1 hora.
    Se auto-invalida si algún Excel fue modificado después del último cache.
    Usa force=True para forzar recarga.
    """
    global _cache_ts
    now = _time.time()
    # Auto-invalidar si: forzado, TTL expirado, o excels más nuevos
    if _cache and not force:
        if (now - _cache_ts) < _CACHE_TTL:
            # Verificar si excels fueron actualizados
            try:
                if _excels_mtime() <= _cache_ts:
                    return _cache
            except Exception:
                return _cache
        print("[analitica] Cache expirado o excels actualizados, recargando…")

    # Serializar la carga pesada: si otro hilo ya la está haciendo,
    # este espera y luego encuentra el caché listo.
    with _cargar_lock:
        # Re-chequear bajo el lock: otro hilo pudo haber llenado el caché
        now = _time.time()
        if _cache and not force and (now - _cache_ts) < _CACHE_TTL:
            try:
                if _excels_mtime() <= _cache_ts:
                    return _cache
            except Exception:
                return _cache

        carpeta = _carpeta()
        t0 = _time.time()
        df_todos = gp.cargar_todos_excels(carpeta)
        bodega, farm_stock_ult, farm_todo = gp.cargar_sap_completo(carpeta)
        # Universo = PDVs farmacias con venta>0 O stock>0 en TODOS los meses
        # Usar CODIGOPDV (código numérico único) para evitar duplicados por variación de nombre
        df_farm_todos = df_todos[df_todos["UNIDAD"] == "FARMACIAS"] if not df_todos.empty else pd.DataFrame()
        _id_col = "CODIGOPDV" if ("CODIGOPDV" in df_farm_todos.columns if not df_farm_todos.empty else False) else "POS"
        if not df_farm_todos.empty and _id_col in df_farm_todos.columns:
            _pv = set(df_farm_todos[df_farm_todos["VENTA NETA RECUPERO"] > 0][_id_col].dropna().unique()) if "VENTA NETA RECUPERO" in df_farm_todos.columns else set()
            _ps = set(df_farm_todos[df_farm_todos["STOCK"] > 0][_id_col].dropna().unique()) if "STOCK" in df_farm_todos.columns else set()
            universo = len(_pv | _ps) if (_pv or _ps) else gp.calcular_universo_pdv(carpeta)
        else:
            universo = gp.calcular_universo_pdv(carpeta)
        stock_por_mes = gp.cargar_stock_por_mes(carpeta)
        ultimo_dia, dias_mes, mes_completo = gp.detectar_ultimo_dia_y_proyeccion(carpeta)

        # Pre-cachear SAP DataFrame para evitar re-lectura de Excel en cada request
        df_sap_cached = None
        try:
            sap_path = gp.detectar_archivo_sap(carpeta)
            if sap_path:
                df_sap_cached = pd.read_excel(sap_path)
        except Exception as e:
            print(f"[analitica] WARN: No se pudo pre-cachear SAP ({e}), se usará df_todos como fallback")

        elapsed = round(_time.time() - t0, 1)
        print(f"[analitica] Data cargada en {elapsed}s — {len(df_todos)} filas")

        _cache.update({
            "df_todos": df_todos,
            "bodega": bodega,
            "farm_stock_ult": farm_stock_ult,
            "farm_todo": farm_todo,
            "universo_pdv": universo,
            "stock_por_mes": stock_por_mes,
            "ultimo_dia_venta": ultimo_dia,
            "dias_mes": dias_mes,
            "mes_completo": mes_completo,
            "df_sap": df_sap_cached,
        })
        _cache_ts = _time.time()
        return _cache


def invalidar_cache():
    global _cache_ts, _cache_pareto, _cache_pareto_ts, _cache_pareto_grupo
    _cache.clear()
    _cache_ts = 0
    _cache_pareto = None
    _cache_pareto_ts = 0
    _cache_pareto_grupo = {}


# ══════════════════════════════════════════════════════════════
# Mapeo GRUPOPDV → nombre Genomma (agrupado)
# ══════════════════════════════════════════════════════════════

_GRUPO_DISPLAY = {
    "Cafa Mostrador":     "Cruz Azul Mostrador",
    "Cafi Mostrador":     "Cruz Azul Mostrador",
    "Cofa Mostrador":     "Cruz Azul Mostrador",
    "Cafa Autoservicio":  "Cruz Azul Autoservicios",
    "Cafi Autoservicio":  "Cruz Azul Autoservicios",
    "Cofa Autoservicio":  "Cruz Azul Autoservicios",
    # Los demás se muestran tal cual: Pharmacys, Dromayor, etc.
}

def _grupo_display(raw: str) -> str:
    """Convierte nombre interno DIFARE → nombre Genomma."""
    return _GRUPO_DISPLAY.get(raw, raw)

def _grupo_raw_values(display_names: list[str]) -> list[str]:
    """Dado nombres Genomma, devuelve TODOS los valores raw que matchean."""
    # Build reverse map: display_name → [raw1, raw2, ...]
    rev: dict[str, list[str]] = {}
    for raw, disp in _GRUPO_DISPLAY.items():
        rev.setdefault(disp, []).append(raw)
    result = []
    for name in display_names:
        if name in rev:
            result.extend(rev[name])
        else:
            result.append(name)  # Pharmacys, Dromayor, etc. → sin mapeo
    return result


# ══════════════════════════════════════════════════════════════
# KPIs principales (alimenta /api/kpis)
# ══════════════════════════════════════════════════════════════

def _aplicar_filtros_df(df, marca=None, canal=None, grupos=None, productos=None):
    """Aplica filtros comunes a un DataFrame de ventas.

    `marca` puede ser:
      - None → sin filtro
      - str  → match por substring case-insensitive (compatibilidad legacy)
      - list → match exacto contra cualquiera de los nombres
    """
    if marca:
        if isinstance(marca, (list, tuple, set)):
            marcas_set = set(str(m).strip() for m in marca if m)
            if marcas_set:
                df = df[df["MARCA"].astype(str).isin(marcas_set)]
        else:
            df = df[df["MARCA"].astype(str).str.contains(str(marca), case=False, na=False)]
    if canal:
        df = df[df["UNIDAD"] == canal]
    if grupos and "GRUPOPDV" in df.columns:
        # Expandir nombres Genomma → valores raw DIFARE
        raw_vals = _grupo_raw_values(grupos)
        df = df[df["GRUPOPDV"].isin(raw_vals)]
    if productos:
        df = df[df["PRODUCTO"].isin(productos)]
    return df


def kpis_dashboard(marca: str | None = None, canal: str | None = None,
                   grupos: list | None = None, productos: list | None = None) -> dict:
    d = cargar_data()
    df = d["df_todos"]
    if df.empty:
        return {"error": "no hay data"}

    df = _aplicar_filtros_df(df, marca=marca, canal=canal, grupos=grupos, productos=productos)

    farm = df[df["UNIDAD"] == "FARMACIAS"]
    dist = df[df["UNIDAD"] == "DISTRIBUCION DIFARE"]

    venta_farm = float(farm["VENTA NETA RECUPERO"].sum())
    venta_dist = float(dist["VENTA NETA RECUPERO"].sum())
    venta_total = venta_farm + venta_dist

    return {
        "venta_total": venta_total,
        "venta_farmacias": venta_farm,
        "venta_distribucion": venta_dist,
        "universo_pdv": d["universo_pdv"],
        "ultimo_dia_venta": d["ultimo_dia_venta"],
        "dias_mes": d["dias_mes"],
        "mes_completo": d["mes_completo"],
        "stock_por_mes": d["stock_por_mes"],
    }


def filtros_disponibles(marca=None) -> dict:
    """Retorna las opciones de filtros disponibles: marcas, grupos, productos (cascadeados por marca).

    `marca` acepta None, str (substring match, legacy) o list (match exacto multi).
    """
    d = cargar_data()
    df = d["df_todos"]
    farm_todo = d.get("farm_todo")

    # Marcas de todos los datos
    marcas = sorted(df["MARCA"].dropna().unique().tolist()) if "MARCA" in df.columns else []

    # Canales (unidades de negocio)
    canales = []
    if "UNIDAD" in df.columns:
        canales = [u for u in ["FARMACIAS", "DISTRIBUCION DIFARE"]
                   if u in df["UNIDAD"].unique()]

    # Grupos PDV desde farm_todo (farmacias) — mapeados a nombres Genomma
    grupos = []
    if farm_todo is not None and not farm_todo.empty:
        if "GRUPOPDV" in farm_todo.columns:
            raw_grupos = farm_todo["GRUPOPDV"].dropna().unique().tolist()
            grupos = sorted(set(_grupo_display(g) for g in raw_grupos))

    # Productos — cascadeados por marca(s) si se pasa
    df_prod = df
    if marca:
        if isinstance(marca, (list, tuple, set)):
            marcas_set = set(str(m).strip() for m in marca if m)
            if marcas_set:
                df_prod = df_prod[df_prod["MARCA"].astype(str).isin(marcas_set)]
        else:
            df_prod = df_prod[df_prod["MARCA"].astype(str).str.contains(str(marca), case=False, na=False)]
    productos = sorted(df_prod["PRODUCTO"].dropna().unique().tolist()) if "PRODUCTO" in df_prod.columns else []

    # Mapeo marca → productos para filtrado dinámico en frontend
    productos_por_marca = {}
    if "MARCA" in df.columns and "PRODUCTO" in df.columns:
        for m in marcas:
            prods_m = sorted(df[df["MARCA"] == m]["PRODUCTO"].dropna().unique().tolist())
            if prods_m:
                productos_por_marca[m] = prods_m

    return {
        "marcas": marcas,
        "canales": canales,
        "grupos": grupos,
        "productos": productos,
        "productos_por_marca": productos_por_marca,
    }


# ══════════════════════════════════════════════════════════════
# 1) Tendencia por marca (pregunta KAM #1)
# ══════════════════════════════════════════════════════════════

def tendencia_marca(unidad_negocio: str | None = None,
                    comparar_yoy: bool = False) -> list[dict]:
    """
    unidad_negocio: 'FARMACIAS' | 'DIFARE S.A.' | None (total)
    comparar_yoy: en Fase 1 siempre False (no hay 2025 cargado todavía).
    Devuelve [{marca, mes, venta}, ...]
    """
    d = cargar_data()
    df = d["df_todos"]
    if df.empty:
        return []
    if unidad_negocio:
        df = df[df["UNIDAD"] == unidad_negocio]
    g = df.groupby(["MARCA", "MES"], dropna=True)["VENTA NETA RECUPERO"].sum().reset_index()
    g = g.rename(columns={"VENTA NETA RECUPERO": "venta"})
    return g.to_dict(orient="records")


def _mes_num(mes_raw) -> int | None:
    """Convierte 'MES' (formatos: '2026-01', '01', 1, '2026-1') a int 1-12."""
    if mes_raw is None:
        return None
    s = str(mes_raw).strip()
    if s in ("", "desconocido", "nan"):
        return None
    # Formato "YYYY-MM" o "YYYY-M"
    if "-" in s:
        try:
            return int(s.split("-")[1])
        except (ValueError, IndexError):
            return None
    # Formato numérico directo "1", "01", etc.
    try:
        n = int(float(s))
        return n if 1 <= n <= 12 else None
    except ValueError:
        return None


def venta_por_canal_mes(marca: str | None = None, canal: str | None = None,
                        grupos: list | None = None, productos: list | None = None) -> list[dict]:
    """
    Devuelve la venta mensual desglosada por canal.
    Para el mes en curso (incompleto) añade campos de proyección.
    """
    d = cargar_data()
    df = d["df_todos"]
    if df.empty:
        return []
    # Solo canales de venta real (excluye 'DIFARE S.A.' que es bodega)
    df = df[df["UNIDAD"].isin(["FARMACIAS", "DISTRIBUCION DIFARE"])]
    df = _aplicar_filtros_df(df, marca=marca, canal=canal, grupos=grupos, productos=productos)

    g = (df.groupby(["MES", "UNIDAD"], dropna=True)["VENTA NETA RECUPERO"]
           .sum().reset_index())
    piv = g.pivot(index="MES", columns="UNIDAD", values="VENTA NETA RECUPERO").fillna(0)
    piv = piv.reset_index()

    ultimo_dia = int(d.get("ultimo_dia_venta") or 0)
    dias_mes = int(d.get("dias_mes") or 30)
    mes_completo = bool(d.get("mes_completo"))
    # El mes actual (incompleto) es el más alto presente si mes_completo=False
    meses_raw = [_mes_num(r["MES"]) for _, r in piv.iterrows()]
    meses_validos = [m for m in meses_raw if m is not None]
    mes_actual = max(meses_validos) if meses_validos and not mes_completo else None
    factor = (dias_mes / ultimo_dia) if (ultimo_dia and dias_mes) else 1.0

    out = []
    for _, r in piv.iterrows():
        n = _mes_num(r["MES"])
        if n is None:
            continue
        f = float(r.get("FARMACIAS", 0) or 0)
        dist = float(r.get("DISTRIBUCION DIFARE", 0) or 0)
        fila = {
            "mes": n,
            "farmacias": f,
            "distribucion": dist,
            "total": f + dist,
            "proyectado": False,
        }
        if n == mes_actual and factor > 1.0:
            f_proy = f * factor
            dist_proy = dist * factor
            fila.update({
                "proyectado": True,
                "ultimo_dia": ultimo_dia,
                "dias_mes": dias_mes,
                "farmacias_proy": f_proy,
                "distribucion_proy": dist_proy,
                "total_proy": f_proy + dist_proy,
                # delta = lo que falta por vender para llegar a la proyección
                "farmacias_delta": max(f_proy - f, 0),
                "distribucion_delta": max(dist_proy - dist, 0),
                "total_delta": max((f_proy + dist_proy) - (f + dist), 0),
            })
        out.append(fila)
    out.sort(key=lambda x: x["mes"])
    return out


# ══════════════════════════════════════════════════════════════
# 4) Rankings (pregunta KAM #4)
# ══════════════════════════════════════════════════════════════

def ranking_pdv(canal: str = "FARMACIAS", top_n: int = 50) -> list[dict]:
    d = cargar_data()
    df = d["df_todos"]
    df = df[df["UNIDAD"] == canal]
    if df.empty:
        return []
    # FARMACIAS → key = POS (nombre + código de PDV).
    # DISTRIBUCION DIFARE → key = PROPIETARIO (cliente del canal distributivo).
    if canal == "FARMACIAS":
        key, label_col = "POS", "PROVINCIA"
    else:
        key, label_col = "PROPIETARIO", "GRUPOCLIENTE"

    cols_extra = [c for c in (label_col, "PROVINCIA", "CIUDAD", "GRUPOPDV", "GRUPOCLIENTE")
                  if c in df.columns and c != key]
    g = (df.groupby(key, dropna=False)
           .agg(**{
               "venta": ("VENTA NETA RECUPERO", "sum"),
               **{c: (c, "first") for c in cols_extra}
           })
           .reset_index()
           .rename(columns={key: "cliente"})
           .sort_values("venta", ascending=False)
           .head(top_n))
    return g.to_dict(orient="records")


def pareto_pdv() -> list[dict]:
    """Top farmacias que acumulan el 80% de la venta."""
    d = cargar_data()
    df = d["df_todos"]
    df = df[df["UNIDAD"] == "FARMACIAS"]
    if df.empty:
        return []
    g = (df.groupby("POS")["VENTA NETA RECUPERO"].sum()
           .sort_values(ascending=False).reset_index())
    total = g["VENTA NETA RECUPERO"].sum()
    if total <= 0:
        return []
    g["pct"] = g["VENTA NETA RECUPERO"] / total * 100
    g["pct_acum"] = g["pct"].cumsum()
    return g[g["pct_acum"] <= 80].to_dict(orient="records")


# ══════════════════════════════════════════════════════════════
# 2) Días de inventario y proyección de venta (pregunta KAM #2)
# ══════════════════════════════════════════════════════════════

def dias_inventario(producto: str | None = None, marca: str | None = None,
                    canal: str | None = None, grupos: list | None = None) -> dict:
    """
    DOIS (Días de Inventario) — fórmula idéntica al reporte PDF de agente-excel.

    DOIS = Stock_Valorizado / Venta_Diaria_SAP
    Donde Venta_Diaria_SAP = Venta del archivo SAP (mes actual) / dias transcurridos.

    Importante:
    - Bodega DOIS: stock bodega / venta diaria (FARMACIAS + DISTRIBUCIÓN)
    - PDV DOIS:    stock PDV   / venta diaria (solo FARMACIAS)
    - Total DOIS:  stock total / venta diaria (FARMACIAS + DISTRIBUCIÓN)
    - Usa SOLO venta del SAP (mes actual), NO la venta acumulada de todos los meses
    """
    d = cargar_data()
    bodega = d["bodega"]
    farm_stock = d["farm_stock_ult"]
    ultimo_dia = max(d["ultimo_dia_venta"], 1)
    dias_mes = d.get("dias_mes", 30) or 30

    # ── Venta del SAP (mes actual) ──
    # Usar SAP cacheado en cargar_data() (no re-leer Excel del disco)
    df_sap = d.get("df_sap")

    # Helper: aplicar filtros a un DataFrame
    # venta: todos los filtros | bodega: solo marca/producto | pdv: marca/grupo/producto
    # marca acepta string o lista; producto acepta string o lista.
    def _filtrar_marca(_df):
        if not marca or "MARCA" not in _df.columns:
            return _df
        if isinstance(marca, (list, tuple, set)):
            marcas_up = set(str(m).upper() for m in marca if m)
            if not marcas_up:
                return _df
            return _df[_df["MARCA"].astype(str).str.upper().isin(marcas_up)]
        return _df[_df["MARCA"].astype(str).str.upper() == str(marca).upper()]

    def _filtrar_producto(_df):
        if not producto or "PRODUCTO" not in _df.columns:
            return _df
        if isinstance(producto, (list, tuple, set)):
            prods_set = set(str(p) for p in producto if p)
            if not prods_set:
                return _df
            return _df[_df["PRODUCTO"].astype(str).isin(prods_set)]
        return _df[_df["PRODUCTO"].astype(str).str.contains(str(producto), case=False, na=False)]

    def _filtro_venta(df):
        _df = _filtrar_marca(df)
        if canal and "UNIDAD" in _df.columns:
            _df = _df[_df["UNIDAD"] == canal]
        if grupos and "GRUPOPDV" in _df.columns:
            raw_vals = _grupo_raw_values(grupos)
            _df = _df[_df["GRUPOPDV"].isin(raw_vals)]
        _df = _filtrar_producto(_df)
        return _df

    def _filtro_stock(df, es_bodega=False):
        _df = _filtrar_marca(df)
        # Bodega (DIFARE S.A.) no tiene GRUPOPDV ni canal de farmacia
        if not es_bodega and grupos and "GRUPOPDV" in _df.columns:
            raw_vals = _grupo_raw_values(grupos)
            _df = _df[_df["GRUPOPDV"].isin(raw_vals)]
        _df = _filtrar_producto(_df)
        return _df

    if df_sap is not None:
        df_sap_f = _filtro_venta(df_sap)
        # Venta: si hay filtro de canal, respetar; si no, usar Farm+Dist y Farm
        if canal:
            venta_sap_total = float(df_sap_f["VENTA NETA RECUPERO"].sum())
            # Si filtra Farmacias: farm_dist=farm, farm=farm
            # Si filtra Distribución: farm_dist=dist, farm=0
            venta_sap_farm_dist = venta_sap_total
            venta_sap_farm = venta_sap_total if canal == "FARMACIAS" else 0
        else:
            venta_sap_farm_dist = float(df_sap_f[df_sap_f["UNIDAD"].isin(
                ["FARMACIAS", "DISTRIBUCION DIFARE"])]["VENTA NETA RECUPERO"].sum())
            venta_sap_farm = float(df_sap_f[df_sap_f["UNIDAD"] == "FARMACIAS"]["VENTA NETA RECUPERO"].sum())
    else:
        df_todos = _filtro_venta(d["df_todos"])
        if canal:
            venta_sap_total = float(df_todos["VENTA NETA RECUPERO"].sum())
            venta_sap_farm_dist = venta_sap_total
            venta_sap_farm = venta_sap_total if canal == "FARMACIAS" else 0
        else:
            venta_sap_farm_dist = float(df_todos[df_todos["UNIDAD"].isin(
                ["FARMACIAS", "DISTRIBUCION DIFARE"])]["VENTA NETA RECUPERO"].sum())
            venta_sap_farm = float(df_todos[df_todos["UNIDAD"] == "FARMACIAS"]["VENTA NETA RECUPERO"].sum())

    # Stock: bodega no filtra por canal/grupo; PDV filtra por grupo pero no canal
    bodega = _filtro_stock(bodega, es_bodega=True)
    farm_stock = _filtro_stock(farm_stock, es_bodega=False)

    # Stock VALORIZADO (USD)
    # Si se filtra por canal o grupo SIN producto específico → bodega = 0
    # (la bodega surte ambos canales; sin producto no tiene sentido mostrarla)
    ocultar_bodega = (canal or grupos) and not producto
    stock_bodega_val = 0 if ocultar_bodega else (
        float(bodega["STOCK_VALORIZADO"].sum()) if "STOCK_VALORIZADO" in bodega.columns else 0
    )
    stock_pdv_val = float(farm_stock["STOCK_VALORIZADO"].sum()) if "STOCK_VALORIZADO" in farm_stock.columns else 0
    stock_total_val = stock_bodega_val + stock_pdv_val

    # Venta diaria del SAP (mes actual)
    venta_diaria_farm_dist = venta_sap_farm_dist / ultimo_dia if ultimo_dia else 0
    venta_diaria_farm = venta_sap_farm / ultimo_dia if ultimo_dia else 0

    # DOIS = Stock Valorizado / Venta Diaria
    # Bodega y Total usan venta Farm+Dist; PDV usa solo Farm
    # Si bodega oculta → DOIS bodega = None, Total usa misma venta que PDV
    if ocultar_bodega:
        dois_bodega = None
        dois_pdv = round(stock_pdv_val / venta_diaria_farm, 1) if venta_diaria_farm > 0 else None
        # Total = solo PDV cuando bodega oculta
        dois_total = dois_pdv
    else:
        dois_bodega = round(stock_bodega_val / venta_diaria_farm_dist, 1) if venta_diaria_farm_dist > 0 else None
        dois_pdv = round(stock_pdv_val / venta_diaria_farm, 1) if venta_diaria_farm > 0 else None
        dois_total = round(stock_total_val / venta_diaria_farm_dist, 1) if venta_diaria_farm_dist > 0 else None

    # Clasificación: >30 días = OK, 15-30 = bajo, <15 = crítico
    if dois_total is not None:
        if dois_total > 30:
            estado = "STOCK OK — inventario saludable"
        elif dois_total >= 15:
            estado = "STOCK BAJO — monitorear reabastecimiento"
        else:
            estado = "STOCK CRÍTICO — riesgo de desabasto"
    else:
        estado = "Sin datos suficientes"

    # Fórmula legible para que Claude la muestre
    fmt = lambda v: f"${v:,.2f}"
    formula_bodega = f"DOIS Bodega = {fmt(stock_bodega_val)} / ({fmt(venta_sap_farm_dist)}/{ultimo_dia}) = {fmt(stock_bodega_val)} / {fmt(venta_diaria_farm_dist)} = {dois_bodega} días"
    formula_pdv = f"DOIS PDV = {fmt(stock_pdv_val)} / ({fmt(venta_sap_farm)}/{ultimo_dia}) = {fmt(stock_pdv_val)} / {fmt(venta_diaria_farm)} = {dois_pdv} días"
    formula_total = f"DOIS Total = {fmt(stock_total_val)} / ({fmt(venta_sap_farm_dist)}/{ultimo_dia}) = {fmt(stock_total_val)} / {fmt(venta_diaria_farm_dist)} = {dois_total} días"

    return {
        "stock_bodega_valorizado": round(stock_bodega_val, 2),
        "stock_pdv_valorizado": round(stock_pdv_val, 2),
        "stock_total_valorizado": round(stock_total_val, 2),
        "venta_sap_farm_dist": round(venta_sap_farm_dist, 2),
        "venta_sap_farm": round(venta_sap_farm, 2),
        "venta_diaria_farm_dist": round(venta_diaria_farm_dist, 2),
        "venta_diaria_farm": round(venta_diaria_farm, 2),
        "dias_transcurridos": ultimo_dia,
        "dias_mes": dias_mes,
        "dois_bodega": dois_bodega,
        "dois_pdv": dois_pdv,
        "dois_total": dois_total,
        "estado_inventario": estado,
        "formula_bodega": formula_bodega,
        "formula_pdv": formula_pdv,
        "formula_total": formula_total,
    }


def dois_por_producto(umbral_min: float = 0, umbral_max: float = 9999,
                      top_n: int = 30) -> list[dict]:
    """
    Calcula DOIS individual por cada producto Pareto (80% de la venta).
    Permite filtrar por rango de DOIS: umbral_min <= dois <= umbral_max.
    Retorna lista ordenada por DOIS descendente.
    """
    d = cargar_data()
    bodega = d["bodega"]
    farm_stock = d["farm_stock_ult"]
    ultimo_dia = max(d["ultimo_dia_venta"], 1)

    # Venta SAP del mes actual (usar cache, no re-leer Excel)
    df_sap = d.get("df_sap")
    if df_sap is None:
        df_sap = d["df_todos"]

    # Productos Pareto (80% de la venta en farmacias)
    farm_sap = df_sap[df_sap["UNIDAD"] == "FARMACIAS"]
    venta_prod = farm_sap.groupby(["IDNEPTUNO", "MARCA", "PRODUCTO"]).agg(
        venta=("VENTA NETA RECUPERO", "sum")
    ).reset_index().sort_values("venta", ascending=False)
    total = venta_prod["venta"].sum()
    if total <= 0:
        return []
    venta_prod["pct"] = venta_prod["venta"] / total * 100
    venta_prod["pct_acum"] = venta_prod["pct"].cumsum()
    pareto = venta_prod[venta_prod["pct_acum"] <= 80].copy()
    if pareto.empty:
        pareto = venta_prod.head(20).copy()

    resultados = []
    for _, row in pareto.iterrows():
        idnep = row["IDNEPTUNO"]
        # Stock valorizado en bodega
        bod_prod = bodega[bodega["IDNEPTUNO"] == idnep]
        stk_bod = float(bod_prod["STOCK_VALORIZADO"].sum()) if "STOCK_VALORIZADO" in bodega.columns and not bod_prod.empty else 0
        # Stock valorizado en PDV
        pdv_prod = farm_stock[farm_stock["IDNEPTUNO"] == idnep]
        stk_pdv = float(pdv_prod["STOCK_VALORIZADO"].sum()) if "STOCK_VALORIZADO" in farm_stock.columns and not pdv_prod.empty else 0
        stk_total = stk_bod + stk_pdv
        # Venta diaria del producto
        venta_diaria = row["venta"] / ultimo_dia if ultimo_dia else 0
        # DOIS
        dois = round(stk_total / venta_diaria, 1) if venta_diaria > 0 else 0
        if umbral_min <= dois <= umbral_max:
            resultados.append({
                "IDNEPTUNO": idnep,
                "MARCA": row["MARCA"],
                "PRODUCTO": row["PRODUCTO"],
                "stock_bodega_val": round(stk_bod, 2),
                "stock_pdv_val": round(stk_pdv, 2),
                "stock_total_val": round(stk_total, 2),
                "venta_mes_sap": round(row["venta"], 2),
                "venta_diaria": round(venta_diaria, 2),
                "dois": dois,
            })

    resultados.sort(key=lambda x: x["dois"], reverse=True)
    return resultados[:top_n]


def proyeccion_venta(horizonte_dias: int = 30) -> dict:
    d = cargar_data()
    df = d["df_todos"]
    farm = df[df["UNIDAD"] == "FARMACIAS"]
    venta = float(farm["VENTA NETA RECUPERO"].sum())
    dias = max(d["ultimo_dia_venta"], 1)
    venta_diaria = venta / dias
    return {
        "venta_actual_mes": venta,
        "dias_transcurridos": dias,
        "venta_diaria_promedio": venta_diaria,
        "proyeccion_horizonte": venta_diaria * horizonte_dias,
        "horizonte_dias": horizonte_dias,
    }


# ══════════════════════════════════════════════════════════════
# 3) Vectorización y venta perdida (pregunta KAM #3)
# ══════════════════════════════════════════════════════════════

_cache_pareto = None
_cache_pareto_ts = 0
_cache_pareto_grupo = {}   # {grupo_key: (rows, timestamp)}
_CACHE_GRUPO_TTL = 300     # 5 minutos para grupo

def _calcular_universo_grupo(d: dict, raw_vals: list, farm_todo_f) -> int:
    """Calcula universo de PDVs para un grupo específico usando 3 meses."""
    _id = "CODIGOPDV" if "CODIGOPDV" in farm_todo_f.columns else "POS"
    codigos_grupo = set(farm_todo_f[_id].dropna().unique()) if not farm_todo_f.empty else set()
    df_todos_all = d["df_todos"]
    df_farm_all = df_todos_all[df_todos_all["UNIDAD"] == "FARMACIAS"]
    if not df_farm_all.empty and _id in df_farm_all.columns:
        mask_grupo = df_farm_all[_id].isin(codigos_grupo)
        if "GRUPOPDV" in df_farm_all.columns:
            mask_grupo = mask_grupo | df_farm_all["GRUPOPDV"].isin(raw_vals)
        df_farm_grupo = df_farm_all[mask_grupo]
        _pv = set(df_farm_grupo[df_farm_grupo["VENTA NETA RECUPERO"] > 0][_id].dropna().unique()) if "VENTA NETA RECUPERO" in df_farm_grupo.columns else set()
        _ps = set(df_farm_grupo[df_farm_grupo["STOCK"] > 0][_id].dropna().unique()) if "STOCK" in df_farm_grupo.columns else set()
        return len(_pv | _ps) if (_pv or _ps) else len(codigos_grupo)
    return len(codigos_grupo)


def _calcular_doi_buckets(d: dict, rows: list, df_farm_src=None) -> list:
    """
    Calcula DOI por PDV por producto y agrupa en buckets.
    DOI = Stock_Valorizado / promedio_venta_diaria
    Donde promedio_venta_diaria = promedio(
        venta_mes_actual / dias_transcurridos,
        venta_mes_anterior / 30,
        venta_2meses / 30
    )
    Usa STOCK_VALORIZADO y VENTA NETA RECUPERO (ambos en USD).
    ** Versión vectorizada — sin for-loops por PDV **
    """
    import numpy as np

    df_todos = d["df_todos"]
    farm_stock = d["farm_stock_ult"]
    ultimo_dia = max(d["ultimo_dia_venta"], 1)

    # Defaults
    _zeros = {"DOI_LE20": 0, "DOI_20_30": 0, "DOI_30_60": 0, "DOI_GT60": 0}

    df_farm = df_farm_src if df_farm_src is not None else df_todos[df_todos["UNIDAD"] == "FARMACIAS"]

    if df_farm.empty or "MES" not in df_farm.columns:
        for r in rows:
            r.update(_zeros)
        return rows

    # Identificar los 3 meses
    meses_ord = sorted(df_farm["MES"].dropna().unique())
    mes_actual = meses_ord[-1] if meses_ord else None
    mes_ant = meses_ord[-2] if len(meses_ord) >= 2 else None
    mes_2ant = meses_ord[-3] if len(meses_ord) >= 3 else None

    stock_col = "STOCK_VALORIZADO" if "STOCK_VALORIZADO" in farm_stock.columns else "STOCK"

    # ── 1) Stock valorizado por (IDNEPTUNO, POS) — un solo groupby ──
    if farm_stock.empty:
        for r in rows:
            r.update(_zeros)
        return rows

    stock_gb = farm_stock.groupby(["IDNEPTUNO", "POS"])[stock_col].sum()
    stock_gb = stock_gb[stock_gb > 0]  # solo PDVs con stock positivo
    if stock_gb.empty:
        for r in rows:
            r.update(_zeros)
        return rows

    # ── 2) Ventas por (IDNEPTUNO, POS, MES) — un solo groupby ──
    venta_gb = df_farm.groupby(["IDNEPTUNO", "POS", "MES"])["VENTA NETA RECUPERO"].sum()

    # Extraer IDs de productos en rows para filtrar
    ids_rows = set(r.get("IDNEPTUNO", "") for r in rows)

    # ── 3) Para cada producto, calcular DOI vectorizado ──
    # Pre-indexar ventas por IDNEPTUNO para acceso rápido
    # Convertir a DataFrame para operaciones masivas
    stock_df = stock_gb.reset_index()
    stock_df.columns = ["IDNEPTUNO", "POS", "STOCK_VAL"]

    # Filtrar solo productos relevantes
    stock_df = stock_df[stock_df["IDNEPTUNO"].isin(ids_rows)]
    if stock_df.empty:
        for r in rows:
            r.update(_zeros)
        return rows

    # Preparar ventas por mes en columnas para todos los productos a la vez
    venta_df = venta_gb.reset_index()
    venta_df.columns = ["IDNEPTUNO", "POS", "MES", "VENTA"]
    venta_df = venta_df[venta_df["IDNEPTUNO"].isin(ids_rows)]

    # Pivotar: una fila por (IDNEPTUNO, POS), columnas = meses
    if not venta_df.empty:
        venta_pivot = venta_df.pivot_table(
            index=["IDNEPTUNO", "POS"], columns="MES",
            values="VENTA", aggfunc="sum", fill_value=0
        )
    else:
        venta_pivot = pd.DataFrame()

    # Merge stock con ventas
    merged = stock_df.set_index(["IDNEPTUNO", "POS"])
    if not venta_pivot.empty:
        merged = merged.join(venta_pivot, how="left").fillna(0)

    # Calcular venta diaria para cada mes presente
    # v_actual / ultimo_dia, v_ant / 30, v_2ant / 30 — solo si > 0
    v_act = merged[mes_actual].values if (mes_actual is not None and mes_actual in merged.columns) else np.zeros(len(merged))
    v_ant_arr = merged[mes_ant].values if (mes_ant is not None and mes_ant in merged.columns) else np.zeros(len(merged))
    v_2ant_arr = merged[mes_2ant].values if (mes_2ant is not None and mes_2ant in merged.columns) else np.zeros(len(merged))

    # Convertir a arrays float
    v_act = np.asarray(v_act, dtype=float)
    v_ant_arr = np.asarray(v_ant_arr, dtype=float)
    v_2ant_arr = np.asarray(v_2ant_arr, dtype=float)

    # Diarios por mes (0 si la venta es <= 0)
    d_act = np.where(v_act > 0, v_act / ultimo_dia, 0.0)
    d_ant = np.where(v_ant_arr > 0, v_ant_arr / 30.0, 0.0)
    d_2ant = np.where(v_2ant_arr > 0, v_2ant_arr / 30.0, 0.0)

    # Cantidad de meses con venta > 0 (para promediar)
    n_meses = (d_act > 0).astype(float) + (d_ant > 0).astype(float) + (d_2ant > 0).astype(float)
    suma_diarios = d_act + d_ant + d_2ant

    # Promedio diario (evitar div/0)
    prom_diario = np.where(n_meses > 0, suma_diarios / n_meses, 0.0)

    # DOI = stock / prom_diario  (si prom_diario == 0 → 999)
    stock_vals = merged["STOCK_VAL"].values.astype(float)
    doi_arr = np.where(prom_diario > 0, stock_vals / prom_diario, 999.0)

    # Asignar bucket por PDV
    bucket_le20 = (doi_arr <= 20).astype(int)
    bucket_20_30 = ((doi_arr > 20) & (doi_arr <= 30)).astype(int)
    bucket_30_60 = ((doi_arr > 30) & (doi_arr <= 60)).astype(int)
    bucket_gt60 = (doi_arr > 60).astype(int)

    # Agregar columna IDNEPTUNO para agrupar
    merged_idx = merged.index.get_level_values("IDNEPTUNO")

    # Crear DataFrame con resultados y sumar por producto
    result_df = pd.DataFrame({
        "IDNEPTUNO": merged_idx,
        "DOI_LE20": bucket_le20,
        "DOI_20_30": bucket_20_30,
        "DOI_30_60": bucket_30_60,
        "DOI_GT60": bucket_gt60
    })
    doi_por_prod = result_df.groupby("IDNEPTUNO")[["DOI_LE20", "DOI_20_30", "DOI_30_60", "DOI_GT60"]].sum()
    doi_dict = doi_por_prod.to_dict(orient="index")

    # Asignar a rows
    for r in rows:
        idn = r.get("IDNEPTUNO", "")
        if idn in doi_dict:
            r.update(doi_dict[idn])
        else:
            r.update(_zeros)

    return rows


def oportunidad_vectorizacion(producto=None,
                              top_n: int = 0,
                              grupo=None,
                              marca=None) -> list[dict]:
    """
    Para cada producto activo, calcula presencia, stock buckets y DOI buckets.
    - solo_pareto=False → devuelve TODOS los ítems con flag es_pareto
    - Calcula DOI por PDV: stock / promedio(proy_actual, venta_ant, venta_2ant) * 30
    Si se pasa 'marca' o 'producto', filtra por texto (string o lista).
    Si se pasa 'grupo', recalcula solo para PDVs de ese(esos) grupo(s).
      grupo puede ser string (single) o lista. Con lista se computa la unión
      de PDVs (sin caché por grupo, porque la combinatoria explota).
    Cache de 1 hora para el cálculo pesado (sin filtro de grupo).
    """
    global _cache_pareto, _cache_pareto_ts, _cache_pareto_grupo

    d = cargar_data()
    now = _time.time()

    # Normalizar grupo a lista de strings si viene con contenido
    grupos_list = None
    if grupo:
        if isinstance(grupo, (list, tuple, set)):
            grupos_list = [str(g).strip() for g in grupo if g and str(g).strip()]
        else:
            grupos_list = [str(grupo).strip()]
        grupos_list = grupos_list or None

    # Si hay filtro de grupo, usar caché solo si es exactamente 1 grupo
    if grupos_list:
        # Caché solo para el caso de 1 grupo; multi-grupo recalcula
        if len(grupos_list) == 1:
            grupo_key = grupos_list[0].lower()
            cached = _cache_pareto_grupo.get(grupo_key)
        else:
            grupo_key = None
            cached = None
        if cached and (now - cached[1]) < _CACHE_GRUPO_TTL:
            rows = [dict(r) for r in cached[0]]
            print(f"[vectorización] TP grupo={grupos_list[0]} desde caché — {len(rows)} productos")
        else:
            raw_vals = _grupo_raw_values(grupos_list)
            df_todos_f = d["df_todos"]
            farm_stock_f = d["farm_stock_ult"]
            farm_todo_f = d["farm_todo"]

            # Filtrar sin .copy() — solo vistas (más rápido)
            if "GRUPOPDV" in df_todos_f.columns:
                df_todos_f = df_todos_f[
                    (df_todos_f["UNIDAD"] != "FARMACIAS") |
                    (df_todos_f["GRUPOPDV"].isin(raw_vals))
                ]
            if "GRUPOPDV" in farm_stock_f.columns:
                farm_stock_f = farm_stock_f[farm_stock_f["GRUPOPDV"].isin(raw_vals)]
            if "GRUPOPDV" in farm_todo_f.columns:
                farm_todo_f = farm_todo_f[farm_todo_f["GRUPOPDV"].isin(raw_vals)]

            universo_f = _calcular_universo_grupo(d, raw_vals, farm_todo_f)

            t0 = _time.time()
            pareto = gp.calcular_pareto_farmacias(
                df_todos_f, farm_stock_f, farm_todo_f, universo_f, solo_pareto=False
            )
            if isinstance(pareto, pd.DataFrame):
                rows = pareto.to_dict(orient="records")
            elif isinstance(pareto, list):
                rows = pareto
            else:
                rows = []

            # Calcular DOI con datos filtrados por grupo
            df_farm_grupo = df_todos_f[df_todos_f["UNIDAD"] == "FARMACIAS"]
            d_grupo = {**d, "farm_stock_ult": farm_stock_f}
            rows = _calcular_doi_buckets(d_grupo, rows, df_farm_src=df_farm_grupo)

            elapsed = round(_time.time() - t0, 1)
            print(f"[vectorización] TP grupo(s)={grupos_list} calculado en {elapsed}s — {len(rows)} productos")

            # Guardar en caché solo para el caso de 1 grupo
            if grupo_key:
                _cache_pareto_grupo[grupo_key] = (rows, now)
                # Limpiar cachés viejos (máx 10 grupos)
                if len(_cache_pareto_grupo) > 10:
                    oldest = min(_cache_pareto_grupo, key=lambda k: _cache_pareto_grupo[k][1])
                    del _cache_pareto_grupo[oldest]

            rows = [dict(r) for r in rows]  # deep copy para no mutar caché
    else:
        # Sin grupo: cachear el cálculo pesado del pareto completo
        if _cache_pareto is None or (now - _cache_pareto_ts) > _CACHE_TTL:
            t0 = _time.time()
            pareto = gp.calcular_pareto_farmacias(
                d["df_todos"], d["farm_stock_ult"], d["farm_todo"], d["universo_pdv"],
                solo_pareto=False
            )
            if isinstance(pareto, pd.DataFrame):
                _cache_pareto = pareto.to_dict(orient="records")
            elif isinstance(pareto, list):
                _cache_pareto = pareto
            else:
                _cache_pareto = []

            # Calcular DOI para todos
            _cache_pareto = _calcular_doi_buckets(d, _cache_pareto)

            _cache_pareto_ts = now
            elapsed = round(_time.time() - t0, 1)
            print(f"[vectorización] TP calculado en {elapsed}s — {len(_cache_pareto)} productos")

        rows = [dict(r) for r in _cache_pareto]  # deep copy

    # Filtros de texto: marca y/o producto (ambos aceptan string o lista)
    if marca:
        if isinstance(marca, (list, tuple, set)):
            marca_set = set(str(m).strip().lower() for m in marca if m)
            if marca_set:
                rows = [r for r in rows if str(r.get("MARCA", "")).strip().lower() in marca_set]
        else:
            m = str(marca).lower()
            rows = [r for r in rows if m in str(r.get("MARCA", "")).lower()]
    if producto:
        if isinstance(producto, (list, tuple, set)):
            prod_set = set(str(p).lower() for p in producto if p)
            rows = [r for r in rows if str(r.get("PRODUCTO", "")).lower() in prod_set]
        else:
            p = str(producto).lower()
            rows = [r for r in rows if p in str(r.get("PRODUCTO", "")).lower()]

    if top_n > 0:
        rows = rows[:top_n]

    return rows


# ══════════════════════════════════════════════════════════════
# 5) Sugerido min/max de stock por grupo de farmacias (pregunta KAM #5)
# ══════════════════════════════════════════════════════════════

def sugerido_stock(grupo_farmacia: str | None = None,
                   productos_top: int = 20,
                   lead_time: int = LEAD_TIME_DIAS,
                   buffer: int = BUFFER_DIAS) -> list[dict]:
    """
    Sugerido por (grupo_farmacia, producto):
      - velocidad = venta_unidades_mes / dias_transcurridos
      - min = velocidad × (lead_time + buffer)
      - max = min × 1.5
    grupo_farmacia: nombre o None (todos).
    """
    d = cargar_data()
    df = d["df_todos"]
    farm = df[df["UNIDAD"] == "FARMACIAS"].copy()
    dias = max(d["ultimo_dia_venta"], 1)

    if grupo_farmacia and "GRUPO" in farm.columns:
        farm = farm[farm["GRUPO"].astype(str).str.contains(grupo_farmacia, case=False, na=False)]

    # Identificar productos top por venta
    top_prods = (farm.groupby("PRODUCTO")["VENTA NETA RECUPERO"].sum()
                     .sort_values(ascending=False).head(productos_top).index.tolist())
    farm_top = farm[farm["PRODUCTO"].isin(top_prods)]

    grupo_col = "GRUPO" if "GRUPO" in farm_top.columns else "ESTABLECIMIENTO"
    g = (farm_top.groupby([grupo_col, "PRODUCTO"])
                 .agg(venta_unidades=("CANTIDAD", "sum") if "CANTIDAD" in farm_top.columns
                                     else ("VENTA NETA RECUPERO", "sum"))
                 .reset_index())
    g["velocidad_diaria"] = g["venta_unidades"] / dias
    g["min_sugerido"] = (g["velocidad_diaria"] * (lead_time + buffer)).round().astype(int)
    g["max_sugerido"] = (g["min_sugerido"] * 1.5).round().astype(int)
    g["lead_time"] = lead_time
    g["buffer"] = buffer
    return g.to_dict(orient="records")


# ══════════════════════════════════════════════════════════════
# 6) Export Excel de vectorización (pregunta KAM #6)
# ══════════════════════════════════════════════════════════════

def exportar_vectorizacion_excel(producto: str | list = "", ruta_salida: str = "",
                                 marca: str = "", grupo: str = "",
                                 tipos_pdv: list | None = None) -> str:
    """
    Genera el Informe de Vectorización semanal.
    Una pestaña por MARCA con los productos activos.

    tipos_pdv controla qué PDVs se incluyen (lista de strings):
      - "sin_vectorizar" → PDVs sin presencia histórica (sugerido=2)
      - "stock_0"        → PDVs con stock=0 (sugerido=2)
      - "doi_lte20"      → PDVs con DOI ≤20 días (sugerido dinámico para llegar a 30-60 días)
      - "doi_20_30"      → PDVs con DOI 20-30 días
      - "doi_30_60"      → PDVs con DOI 30-60 días
      - "doi_gt60"       → PDVs con DOI >60 días
    Default (None o vacío): sin_vectorizar + stock_0 + doi_lte20

    Columnas: GRUPOPDV, CODIGOPDV, POS, IDNEPTUNO, IDDIFARE, PRODUCTO, DOIS, STOCK UNIDADES, SUGERIDO

    Columna DOIS muestra:
      - "No aplica" para sin vectorizar
      - 0 para stock=0
      - días reales de inventario para los demás
    """
    # Defaults
    if not tipos_pdv:
        tipos_pdv = ["sin_vectorizar", "stock_0", "doi_lte20"]

    d = cargar_data()
    farm_todo = d["farm_todo"].copy()
    farm_stock = d["farm_stock_ult"].copy()
    df_todos = d["df_todos"]
    ultimo_dia = max(d["ultimo_dia_venta"], 1)

    # Filtrar por grupo si aplica
    if grupo:
        raw_vals = _grupo_raw_values([grupo])
        if "GRUPOPDV" in farm_todo.columns:
            farm_todo = farm_todo[farm_todo["GRUPOPDV"].isin(raw_vals)]
        if "GRUPOPDV" in farm_stock.columns:
            farm_stock = farm_stock[farm_stock["GRUPOPDV"].isin(raw_vals)]

    # Obtener productos (filtrados por marca/grupo/producto si se pasan)
    pareto_rows = oportunidad_vectorizacion(
        producto=producto if producto else None,
        marca=marca if marca else None,
        grupo=grupo if grupo else None,
    )
    if not pareto_rows:
        raise ValueError("No se encontraron productos para generar el informe")

    # Universo de PDV activos
    universo_pdv = set(farm_todo["POS"].dropna().unique())

    # Info de cada PDV: GRUPOPDV, CODIGOPDV, etc.
    pdv_info = (farm_todo.groupby("POS")
                         .agg(GRUPOPDV=("GRUPOPDV", "first"),
                              CODIGOPDV=("CODIGOPDV", "first"))
                         .reset_index())
    pdv_info_map = {r["POS"]: r for _, r in pdv_info.iterrows()}

    # Preparar datos de venta para cálculo DOI
    df_farm = df_todos[df_todos["UNIDAD"] == "FARMACIAS"] if not df_todos.empty else pd.DataFrame()
    if grupo and not df_farm.empty and "GRUPOPDV" in df_farm.columns:
        raw_vals_g = _grupo_raw_values([grupo])
        df_farm = df_farm[df_farm["GRUPOPDV"].isin(raw_vals_g)]

    meses_ord = sorted(df_farm["MES"].dropna().unique()) if not df_farm.empty and "MES" in df_farm.columns else []
    mes_actual = meses_ord[-1] if meses_ord else None
    mes_ant = meses_ord[-2] if len(meses_ord) >= 2 else None
    mes_2ant = meses_ord[-3] if len(meses_ord) >= 3 else None

    stock_val_col = "STOCK_VALORIZADO" if "STOCK_VALORIZADO" in farm_stock.columns else "STOCK"

    def _calcular_doi_pdv(pos, idneptuno, prod_farm, prod_stock_df):
        """Calcula DOI en días para un PDV+producto específico.
        Retorna (doi_dias, stock_valorizado, prom_diario_dolar)."""
        # Stock valorizado del PDV
        pdv_st = prod_stock_df[prod_stock_df["POS"] == pos]
        if pdv_st.empty:
            return 0.0, 0.0, 0.0
        stock_val = float(pdv_st[stock_val_col].sum())
        if stock_val <= 0:
            return 0.0, 0.0, 0.0

        # Venta mensual por mes
        pdv_venta = prod_farm[prod_farm["POS"] == pos]
        venta_por_mes = pdv_venta.groupby("MES")["VENTA NETA RECUPERO"].sum() if not pdv_venta.empty else pd.Series(dtype=float)

        v_actual = float(venta_por_mes.get(mes_actual, 0)) if mes_actual else 0
        v_ant = float(venta_por_mes.get(mes_ant, 0)) if mes_ant else 0
        v_2ant = float(venta_por_mes.get(mes_2ant, 0)) if mes_2ant else 0

        diarios = []
        if v_actual > 0 and ultimo_dia > 0:
            diarios.append(v_actual / ultimo_dia)
        if v_ant > 0:
            diarios.append(v_ant / 30)
        if v_2ant > 0:
            diarios.append(v_2ant / 30)

        if diarios:
            prom_diario = sum(diarios) / len(diarios)
            return stock_val / prom_diario, stock_val, prom_diario
        return 999.0, stock_val, 0.0  # stock sin venta → DOI infinito

    def _sugerido_doi(doi_actual, stock_unidades, stock_valorizado, prom_diario_dolar):
        """Calcula unidades sugeridas para llevar el DOI de ≤20 a ~45 días.
        Usa la relación precio_unit = stock_valorizado / stock_unidades
        para convertir la rotación en dólares a rotación en unidades."""
        if stock_unidades <= 0 or stock_valorizado <= 0 or prom_diario_dolar <= 0:
            return 2  # fallback
        precio_unit = stock_valorizado / stock_unidades
        prom_diario_units = prom_diario_dolar / precio_unit
        target_dias = 45  # punto medio entre 30 y 60
        unidades_target = max(int(round(prom_diario_units * target_dias)), 2)
        sugerido = unidades_target - int(stock_unidades)
        return max(sugerido, 1)

    # Clasificar GRUPOPDV para Suerox
    SUEROX_GRUPO_A = {"Cafa Mostrador", "Cafi Mostrador", "Cofa Mostrador",
                      "Bodegas Internas Privadas", "Bodegas Administrativas"}
    SUEROX_GRUPO_B = {"Cafa Autoservicio", "Cafi Autoservicio", "Pharmacys", "Dromayor"}

    def sugerido_suerox_a(stock_status):
        if stock_status == "SIN VECTORIZAR": return 6
        if stock_status == 0: return 6
        return 2

    def sugerido_suerox_b(stock_status):
        if stock_status == "SIN VECTORIZAR": return 12
        if stock_status == 0: return 12
        return 2

    # Agrupar Pareto por marca
    marcas_dict = {}
    for r in pareto_rows:
        m = r.get("MARCA", "Otros")
        marcas_dict.setdefault(m, []).append(r)

    if not ruta_salida:
        import tempfile
        ruta_salida = os.path.join(tempfile.gettempdir(), "vectorizacion_semanal.xlsx")
    os.makedirs(os.path.dirname(ruta_salida) or ".", exist_ok=True)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1B3A6B")
    data_font = Font(name="Arial", size=9)
    gold_font = Font(name="Arial", size=9, bold=True, color="C9A84C")
    red_font = Font(name="Arial", size=9, bold=True, color="FF0000")
    blue_font = Font(name="Arial", size=9, bold=True, color="2E75B6")
    border = Border(
        bottom=Side(style="thin", color="D0D0D0")
    )
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    resumen_rows = []

    include_sin_vect = "sin_vectorizar" in tipos_pdv
    include_stock0 = "stock_0" in tipos_pdv
    include_doi_lte20 = "doi_lte20" in tipos_pdv
    include_doi_20_30 = "doi_20_30" in tipos_pdv
    include_doi_30_60 = "doi_30_60" in tipos_pdv
    include_doi_gt60 = "doi_gt60" in tipos_pdv

    for marca_name, prods in marcas_dict.items():
        sheet_name = marca_name[:31]  # Excel 31 char limit
        ws = wb.create_sheet(title=sheet_name)

        headers = ["GRUPOPDV", "CODIGOPDV", "POS", "IDNEPTUNO", "IDDIFARE",
                    "PRODUCTO", "DOIS", "STOCK UNIDADES", "SUGERIDO"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        row_idx = 2
        total_sugerido = 0
        total_pdv_accionables = 0

        for prod_info in prods:
            idneptuno = prod_info.get("IDNEPTUNO", "")
            producto_nombre = prod_info.get("PRODUCTO", "")
            is_suerox = marca_name.upper() == "SUEROX"

            # DataFrames de stock y presencia para este producto
            prod_stock_df = farm_stock[farm_stock["IDNEPTUNO"] == idneptuno] if farm_stock is not None and not farm_stock.empty else pd.DataFrame()
            prod_presencia_df = farm_todo[farm_todo["IDNEPTUNO"] == idneptuno] if farm_todo is not None and not farm_todo.empty else pd.DataFrame()
            prod_farm = df_farm[df_farm["IDNEPTUNO"] == idneptuno] if not df_farm.empty else pd.DataFrame()

            pdv_con_presencia = set(prod_presencia_df["POS"].dropna().unique())
            pdv_stock_map = {}
            if not prod_stock_df.empty:
                for _, sr in prod_stock_df.iterrows():
                    pos = sr.get("POS")
                    if pd.notna(pos):
                        pdv_stock_map[pos] = sr.get("STOCK", 0) or 0

            pdv_sin_vectorizar = universo_pdv - pdv_con_presencia

            filas_producto = []

            # 1) Sin vectorizar
            if include_sin_vect:
                for pos in sorted(pdv_sin_vectorizar):
                    info = pdv_info_map.get(pos, {})
                    grp = info.get("GRUPOPDV", "") if isinstance(info, dict) else (info["GRUPOPDV"] if "GRUPOPDV" in info else "")
                    if is_suerox:
                        sug = sugerido_suerox_b("SIN VECTORIZAR") if grp in SUEROX_GRUPO_B else sugerido_suerox_a("SIN VECTORIZAR")
                    else:
                        sug = 2
                    filas_producto.append({
                        "GRUPOPDV": grp,
                        "CODIGOPDV": info.get("CODIGOPDV", "") if isinstance(info, dict) else (info["CODIGOPDV"] if "CODIGOPDV" in info else ""),
                        "POS": pos,
                        "IDNEPTUNO": idneptuno,
                        "IDDIFARE": prod_info.get("IDDIFARE", ""),
                        "PRODUCTO": producto_nombre,
                        "DOIS": "No aplica",
                        "STOCK UNIDADES": "SIN VECTORIZAR",
                        "SUGERIDO": sug
                    })

            # 2) PDVs con presencia — clasificar por stock y DOI
            for pos in sorted(pdv_con_presencia):
                stock = pdv_stock_map.get(pos)
                if stock is None:
                    stock = 0

                info = pdv_info_map.get(pos, {})
                grp = info.get("GRUPOPDV", "") if isinstance(info, dict) else (info["GRUPOPDV"] if "GRUPOPDV" in info else "")

                if stock == 0:
                    # Stock = 0
                    if not include_stock0:
                        continue
                    if is_suerox:
                        sug = sugerido_suerox_b(0) if grp in SUEROX_GRUPO_B else sugerido_suerox_a(0)
                    else:
                        sug = 2
                    filas_producto.append({
                        "GRUPOPDV": grp,
                        "CODIGOPDV": info.get("CODIGOPDV", "") if isinstance(info, dict) else (info["CODIGOPDV"] if "CODIGOPDV" in info else ""),
                        "POS": pos,
                        "IDNEPTUNO": idneptuno,
                        "IDDIFARE": prod_info.get("IDDIFARE", ""),
                        "PRODUCTO": producto_nombre,
                        "DOIS": 0,
                        "STOCK UNIDADES": 0,
                        "SUGERIDO": sug
                    })
                else:
                    # Tiene stock > 0 → calcular DOI
                    doi, stock_val, prom_d_dolar = _calcular_doi_pdv(pos, idneptuno, prod_farm, prod_stock_df)
                    doi_rounded = round(doi, 1)

                    # Clasificar en bucket
                    if doi <= 20:
                        if not include_doi_lte20:
                            continue
                        # Sugerido dinámico basado en rotación para llegar a 30-60 días
                        sug = _sugerido_doi(doi, stock, stock_val, prom_d_dolar) if prom_d_dolar > 0 else 2
                    elif doi <= 30:
                        if not include_doi_20_30:
                            continue
                        sug = 0  # No requiere acción urgente, solo informativo
                    elif doi <= 60:
                        if not include_doi_30_60:
                            continue
                        sug = 0
                    else:
                        if not include_doi_gt60:
                            continue
                        sug = 0

                    filas_producto.append({
                        "GRUPOPDV": grp,
                        "CODIGOPDV": info.get("CODIGOPDV", "") if isinstance(info, dict) else (info["CODIGOPDV"] if "CODIGOPDV" in info else ""),
                        "POS": pos,
                        "IDNEPTUNO": idneptuno,
                        "IDDIFARE": prod_info.get("IDDIFARE", ""),
                        "PRODUCTO": producto_nombre,
                        "DOIS": doi_rounded,
                        "STOCK UNIDADES": stock,
                        "SUGERIDO": sug
                    })

            # Escribir filas al Excel
            for fila in filas_producto:
                for c, h in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=c, value=fila[h])
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = left if c <= 6 else center
                    # Resaltar SIN VECTORIZAR en rojo
                    if h == "STOCK UNIDADES" and fila[h] == "SIN VECTORIZAR":
                        cell.font = red_font
                    elif h == "DOIS" and fila[h] == "No aplica":
                        cell.font = blue_font
                    elif h == "SUGERIDO" and fila[h] > 0:
                        cell.font = gold_font
                row_idx += 1
                total_sugerido += fila["SUGERIDO"] if isinstance(fila["SUGERIDO"], (int, float)) else 0
                total_pdv_accionables += 1

        # Ajustar anchos de columna
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 45
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 35
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 18
        ws.column_dimensions["I"].width = 12

        resumen_rows.append({
            "Marca": marca_name,
            "Productos": len(prods),
            "PDVs en reporte": total_pdv_accionables,
            "Total unidades sugeridas": total_sugerido,
        })

    # Hoja Resumen al inicio
    ws_res = wb.create_sheet(title="Resumen", index=0)
    res_headers = ["Marca", "Productos", "PDVs en reporte", "Total unidades sugeridas"]
    for c, h in enumerate(res_headers, 1):
        cell = ws_res.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for i, r in enumerate(resumen_rows, 2):
        for c, h in enumerate(res_headers, 1):
            cell = ws_res.cell(row=i, column=c, value=r[h])
            cell.font = data_font
            cell.alignment = center if c >= 2 else left
    ws_res.column_dimensions["A"].width = 18
    ws_res.column_dimensions["B"].width = 18
    ws_res.column_dimensions["C"].width = 18
    ws_res.column_dimensions["D"].width = 24

    wb.save(ruta_salida)
    return ruta_salida


# ══════════════════════════════════════════════════════════════
# 7) Distribución numérica — clientes atendidos por RUC (canal distributivo)
# ══════════════════════════════════════════════════════════════

def venta_por_canal_farmacia(marca: str | None = None,
                             producto: str | None = None,
                             mes: int | None = None) -> dict:
    """
    Analiza la venta en FARMACIAS desglosada por CANAL (columna 'CANAL'):
    APP, Call Center, Mostrador, Tienda Virtual, etc.

    Clasificación de negocio:
    - Mostrador = Canal PRESENCIAL
    - Todo lo demás (APP, Call Center, Tienda Virtual, etc.) = Canal NO PRESENCIAL

    Devuelve:
    - Desglose por canal con venta, % participación
    - Resumen presencial vs no presencial
    - Evolución mensual si hay varios meses
    """
    d = cargar_data()
    df = d["df_todos"]

    # Solo farmacias
    farm = df[df["UNIDAD"] == "FARMACIAS"].copy()
    if farm.empty:
        return {"error": "No hay datos de farmacias"}

    # Verificar que exista columna CANAL
    if "CANAL" not in farm.columns:
        return {"error": "No se encontró la columna CANAL en los datos"}

    # Filtros opcionales
    if marca:
        farm = farm[farm["MARCA"].astype(str).str.contains(marca, case=False, na=False)]
    if producto:
        farm = farm[farm["PRODUCTO"].astype(str).str.contains(producto, case=False, na=False)]
    if mes:
        farm["_mes_num"] = farm["MES"].apply(_mes_num)
        farm = farm[farm["_mes_num"] == mes]

    if farm.empty:
        return {"error": "No hay datos con los filtros aplicados"}

    venta_col = "VENTA NETA RECUPERO"
    total_venta = float(farm[venta_col].sum())

    # ── Desglose por canal ──
    por_canal = (farm.groupby("CANAL", dropna=False)[venta_col]
                     .sum().reset_index()
                     .rename(columns={venta_col: "venta"})
                     .sort_values("venta", ascending=False))
    por_canal["pct"] = (por_canal["venta"] / total_venta * 100).round(1)
    canales = por_canal.to_dict(orient="records")

    # ── Presencial vs No presencial ──
    venta_presencial = float(farm[farm["CANAL"].astype(str).str.upper().str.strip() == "MOSTRADOR"][venta_col].sum())
    venta_no_presencial = total_venta - venta_presencial
    pct_presencial = round(venta_presencial / total_venta * 100, 1) if total_venta > 0 else 0
    pct_no_presencial = round(100 - pct_presencial, 1)

    resumen_tipo = {
        "presencial_mostrador": round(venta_presencial, 2),
        "no_presencial_otros": round(venta_no_presencial, 2),
        "pct_presencial": pct_presencial,
        "pct_no_presencial": pct_no_presencial,
        "total": round(total_venta, 2),
    }

    # ── Evolución mensual (presencial vs no presencial) ──
    farm["_tipo_canal"] = farm["CANAL"].astype(str).str.upper().str.strip().apply(
        lambda x: "Presencial" if x == "MOSTRADOR" else "No presencial"
    )
    farm["_mes_n"] = farm["MES"].apply(_mes_num)
    evol = (farm.groupby(["_mes_n", "_tipo_canal"], dropna=True)[venta_col]
                .sum().reset_index()
                .rename(columns={venta_col: "venta", "_mes_n": "mes", "_tipo_canal": "tipo"}))
    evol = evol.sort_values(["mes", "tipo"])
    evolucion = evol.to_dict(orient="records")

    return {
        "marca_filtro": marca,
        "producto_filtro": producto,
        "mes_filtro": mes,
        "total_venta_farmacias": round(total_venta, 2),
        "canales": canales,
        "resumen_presencial_vs_no": resumen_tipo,
        "evolucion_mensual": evolucion,
    }


def distribucion_numerica(marca: str | None = None,
                          top_n: int = 20) -> dict:
    """
    Analiza la distribución numérica del canal DISTRIBUCION DIFARE:
    - Clientes únicos atendidos (por RUC) en cada mes
    - Comparativa mensual: ¿cuántos clientes nuevos? ¿cuántos se perdieron?
    - Penetración del portafolio TOP: ¿cuántos productos distintos compra cada cliente?
    Si se pasa 'marca', filtra solo esa marca.
    """
    d = cargar_data()
    df = d["df_todos"]
    dist = df[df["UNIDAD"] == "DISTRIBUCION DIFARE"].copy()

    if dist.empty:
        return {"error": "No hay datos de distribución"}

    # Filtrar por marca si aplica
    if marca:
        dist = dist[dist["MARCA"].astype(str).str.contains(marca, case=False, na=False)]
        if dist.empty:
            return {"error": f"Marca no encontrada en distribución: {marca}"}

    # Identificar columna RUC
    ruc_col = "RUC" if "RUC" in dist.columns else "PROPIETARIO"

    # Asegurar columna MES
    if "MES" not in dist.columns:
        return {"error": "No hay columna MES en los datos"}

    meses = sorted(dist["MES"].dropna().unique().tolist())

    # Clientes por mes
    clientes_por_mes = {}
    for mes in meses:
        mes_data = dist[dist["MES"] == mes]
        rucs = set(mes_data[ruc_col].dropna().unique())
        clientes_por_mes[mes] = rucs

    # Construir resumen mensual
    resumen_meses = []
    prev_rucs = set()
    for mes in meses:
        rucs = clientes_por_mes[mes]
        nuevos = rucs - prev_rucs if prev_rucs else set()
        perdidos = prev_rucs - rucs if prev_rucs else set()
        resumen_meses.append({
            "mes": mes,
            "clientes_atendidos": len(rucs),
            "clientes_nuevos": len(nuevos),
            "clientes_perdidos": len(perdidos),
            "variacion_neta": len(nuevos) - len(perdidos),
        })
        prev_rucs = rucs

    # Portafolio TOP: productos Pareto (80% de venta)
    venta_prod = (dist.groupby("PRODUCTO")["VENTA NETA RECUPERO"].sum()
                      .sort_values(ascending=False))
    total = venta_prod.sum()
    if total > 0:
        venta_prod_pct = (venta_prod / total * 100).cumsum()
        productos_top = venta_prod_pct[venta_prod_pct <= 80].index.tolist()
        if not productos_top:
            productos_top = venta_prod.head(10).index.tolist()
    else:
        productos_top = []

    # Penetración del portafolio TOP por cliente (último mes)
    ultimo_mes = meses[-1] if meses else None
    penetracion = []
    if ultimo_mes and productos_top:
        ult = dist[(dist["MES"] == ultimo_mes)]
        por_cliente = (ult.groupby([ruc_col, "PROPIETARIO"])
                         .agg(
                             productos_comprados=("PRODUCTO", "nunique"),
                             productos_top_comprados=("PRODUCTO", lambda x: len(set(x) & set(productos_top))),
                             venta_total=("VENTA NETA RECUPERO", "sum"),
                         )
                         .reset_index()
                         .sort_values("venta_total", ascending=False))
        por_cliente["pct_portafolio_top"] = (
            por_cliente["productos_top_comprados"] / max(len(productos_top), 1) * 100
        ).round(1)
        penetracion = por_cliente.head(top_n).to_dict(orient="records")

    # Universo total de clientes (histórico)
    todos_rucs = set()
    for rucs in clientes_por_mes.values():
        todos_rucs |= rucs

    return {
        "marca_filtro": marca,
        "total_clientes_historico": len(todos_rucs),
        "productos_top_count": len(productos_top),
        "productos_top": productos_top[:15],  # primeros 15 para contexto
        "resumen_meses": resumen_meses,
        "penetracion_portafolio_ultimo_mes": penetracion,
    }
